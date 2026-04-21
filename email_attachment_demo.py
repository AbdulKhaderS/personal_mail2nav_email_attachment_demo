"""
email_attachment.py  (Mail2Nav — v3)
========================================================================
Mail2Nav: From Inbox to Navision — Automatically.

PURPOSE:
    Monitors your Outlook inbox for new emails ON DEMAND (not automatically).
    When an attachment filename contains a known alias (e.g. "CITY STAR"),
    it renames it to the correct supplier code (e.g. "CCC") and
    saves it into your Operations or CHANGE_PRICE folder on the Desktop.

    CHANGE PRICE files are automatically cleaned before saving:
      - First "Mandatory..." header row is removed
      - Columns removed: Buyer, Arabic Description, Vendor Number
      - Only price-relevant columns are kept

    SPECIAL SENDERS (Amro, Aseel, Hadeel, Zeina, Hussain) use smart detection:
      1. Email body/subject text → keyword scan
      2. Excel columns           → which columns have data decides type
      3. Item NO prefix          → filename like "CHANGE PRICE AF AMRO 07-04-2026"

    COMPLETION REPLY:
      After importing into Navision, click "Done — Send Reply" to open a
      pre-filled Outlook reply to the buyer (CC: ayman.s@example.com).

HOW TO RUN:
    python email_attachment.py
    → A control panel window appears. Click "Process New Emails" when ready.

REQUIREMENTS:
    pip install pywin32 openpyxl

CONFIGURATION:
    Edit the sections marked  <-- CONFIGURE THIS

NOTES:
      One thing to check: If Amro ever adds a new column you want to keep, just add its header name to the KEEP_HEADERS set inside clean_amro_change_price_columns().
"""

import re
import os
import time
import shutil
import logging
import threading
import tkinter as tk
from tkinter import font as tkfont
from datetime import datetime

import win32com.client

# ──────────────────────────────────────────────────────────────
#  PATHS  <-- CONFIGURE THIS
# ──────────────────────────────────────────────────────────────
OPERATIONS_FOLDER   = r"C:\Users\abdul\Desktop\Operations"
CHANGE_PRICE_FOLDER = r"C:\Users\abdul\Desktop\CHANGE_PRICE"
LOG_FILE            = r"C:\Users\abdul\Desktop\email_fixer_log.txt"
BLOCK_TXT_FILE      = r"C:\Users\abdul\Desktop\Block Purc&Disc&Promo&Basicitem.txt"
DMG_TXT_FILE        = r"C:\Users\abdul\Desktop\DMG.txt"

# ── Completion reply settings ──────────────────────────────────
REPLY_CC          = "ayman.s@example.com"
REPLY_SIGNATURE   = "\n\nBest regards,\nAbdul Khader"   # appended to every reply

# ── Session store: tracks every processed email this session ──
# Each entry: { "entry_id", "sender_email", "sender_name",
#               "subject", "saved_filename", "req_type",
#               "replied": bool }
_processed_jobs: list = []

# ──────────────────────────────────────────────────────────────
#  CHANGE PRICE CLEANUP SETTINGS
# ──────────────────────────────────────────────────────────────
#  Columns listed here will be REMOVED from CHANGE PRICE files.
#  Names are matched case-insensitively and by partial match.
# ──────────────────────────────────────────────────────────────
COLUMNS_TO_REMOVE = [
    "BUYER",
    "ARABIC DESCRIPTION",
    "VENDOR NUMBER",
    "VENDER NUMBER",      # common typo variant
    "VENDOR NO",
]

# ──────────────────────────────────────────────────────────────
#  ALIAS MAP  <-- ADD MORE ALIASES HERE ANYTIME
#  Longer / more-specific phrases MUST come before shorter ones.
# ──────────────────────────────────────────────────────────────
ALIAS_MAP = {
    "CITY STAR":             "CCC",
    "CITYSTAR":              "CCC",
    "CITY-STAR":             "CCC",
    "C STAR":                "CCC",

    "CHANGE PRICE & COST":   "CHANGE PRICE",
    "CHANGE PRICE AND COST": "CHANGE PRICE",
    "CHANGE-PRICE & COST":   "CHANGE PRICE",
    "CHANGE PRICE":          "CHANGE PRICE",
    "CHANGEPRICE":           "CHANGE PRICE",
    "CHANGE-PRICE":          "CHANGE PRICE",
    "PRICE CHANGE":          "CHANGE PRICE",
    "PRICECHANGE":           "CHANGE PRICE",
    "REDUCE PRICE":          "CHANGE PRICE",

}

# ──────────────────────────────────────────────────────────────
#  SENDER NAME MAP  <-- ADD MORE SENDERS HERE
# ──────────────────────────────────────────────────────────────
SENDER_NAME_MAP = {
    "hala.r@example.com" :  "HALA",
    "Faisal.m@example.com" : "FAISAL",
    "esraa.k@example.com":  "ISRAA"
    # "john.d@supplier.com": "JOHN",
}

# ──────────────────────────────────────────────────────────────
#  SPECIAL SENDERS
# ──────────────────────────────────────────────────────────────
SPECIAL_SENDERS = {
    "amro.s@example.com":   {"label": "AMRO",   "smart_amro":  True},
    "aseel.d@example.com":  {"label": "ASEEL",  "smart_aseel": True},
    "hadeel.c@example.com": {"label": "HADEEL"},
    "zaina.n@example.com":  {"label": "ZEINA",  "smart_zaina": True},
    "ahmad.w@example.com":  {"label": "AHMAD",  "smart_ahmad": True},
}

# ──────────────────────────────────────────────────────────────
#  KEYWORD DETECTION LIST
# ──────────────────────────────────────────────────────────────
KEYWORD_DETECTION = [
    ("CHANGE PRICE & COST",    "CHANGE PRICE",   CHANGE_PRICE_FOLDER),
    ("CHANGE PRICE AND COST",  "CHANGE PRICE",   CHANGE_PRICE_FOLDER),
    ("CHANGE PRICE ALL SHOPS", "CHANGE PRICE",   CHANGE_PRICE_FOLDER),
    ("CHANGE COST&PRICE",      "CHANGE PRICE",   CHANGE_PRICE_FOLDER),
    ("CHANGE COST & PRICE",    "CHANGE PRICE",   CHANGE_PRICE_FOLDER),
    ("CHANGE COST AND PRICE",  "CHANGE PRICE",   CHANGE_PRICE_FOLDER),
    ("CHANGE COST",            "CHANGE PRICE",   CHANGE_PRICE_FOLDER),
    ("CHANGE PRICE",           "CHANGE PRICE",   CHANGE_PRICE_FOLDER),
    ("CHANGE-PRICE",           "CHANGE PRICE",   CHANGE_PRICE_FOLDER),
    ("PRICE CHANGE",           "CHANGE PRICE",   CHANGE_PRICE_FOLDER),
    ("PRICE UPDATE",           "CHANGE PRICE",   CHANGE_PRICE_FOLDER),
    ("PRICE REVISION",         "CHANGE PRICE",   CHANGE_PRICE_FOLDER),
    ("NEW PRICE",              "CHANGE PRICE",   CHANGE_PRICE_FOLDER),
    ("UPDATED PRICE",          "CHANGE PRICE",   CHANGE_PRICE_FOLDER),
    ("PRICE LIST",             "CHANGE PRICE",   CHANGE_PRICE_FOLDER),
    ("REDUCE PRICE",           "CHANGE PRICE",   CHANGE_PRICE_FOLDER),
    ("NEW ITEM",               "NEW ITEMS",      OPERATIONS_FOLDER),
    ("NEW ITEMS",              "NEW ITEMS",      OPERATIONS_FOLDER),
    ("NEW CREATION",           "NEW ITEMS",      OPERATIONS_FOLDER),   # e.g. Ajmal filename
    ("NEW PRODUCT",            "NEW ITEMS",      OPERATIONS_FOLDER),
    ("NEW PRODUCTS",           "NEW ITEMS",      OPERATIONS_FOLDER),
    ("ADD ITEM",               "NEW ITEMS",      OPERATIONS_FOLDER),
    ("ADD ITEMS",              "NEW ITEMS",      OPERATIONS_FOLDER),
    ("IMPORT",                 "NEW ITEMS",      OPERATIONS_FOLDER),   # Aseel uses "IMPORT"
    ("BLOCK PURCHASE",         "BLOCK PURCHASE", OPERATIONS_FOLDER),
    ("BLOCK",                  "BLOCK PURCHASE", OPERATIONS_FOLDER),
    ("STOP ORDER",             "BLOCK PURCHASE", OPERATIONS_FOLDER),
    ("STOP PURCHASE",          "BLOCK PURCHASE", OPERATIONS_FOLDER),
    ("TRANSFER",               "TRANSFER",       OPERATIONS_FOLDER),
    ("PURCHASE ORDER",         "ORDER",          OPERATIONS_FOLDER),
    ("ORDER",                  "ORDER",          OPERATIONS_FOLDER),
]

# ──────────────────────────────────────────────────────────────
#  Amro template column fingerprints
# ──────────────────────────────────────────────────────────────
AMRO_NEW_ITEM_COLS     = {
    "ENGLISH DESCRIPTION", "BASE UOM", "ITEM CATEGORY",
    "SUB CATEGORY", "SUB-SUB CATEGORY", "ARABIC DESCRIPTION",
    "VENDER NUMBER", "VENDOR ITEMS", "BARCODE",
}
AMRO_CHANGE_PRICE_COLS = {"OLD RSP", "NEW RSP", "OLD COST", "NEW COST"}

# ──────────────────────────────────────────────────────────────
#  BUYER CODE MAP  <-- ADD MORE CODES HERE ANYTIME
#  Maps numeric item-number prefixes to a buyer display name.
#  The buyer name is appended at the END of the saved filename.
#
#  Example item numbers:  W-30044  →  letter prefix = "W"
#                         101-5500 →  numeric prefix = "101" → "IHAB"
#  Result filename:  CHANGE PRICE W IHAB AMRO 08-04-2026.xlsx
#
#  Add more entries as needed:
#    "202": "SARA",
#    "305": "KHALID",
# ──────────────────────────────────────────────────────────────
BUYER_CODE_MAP = {
    "101": "IHAB",
    "106": "HALA",
}

# ──────────────────────────────────────────────────────────────
#  KHOZEMA CONFIG  <-- Fix 1 & 5
#  khozema.b@example.com sends NEW ITEMS files with a
#  DIV or DIVISION column.  The value in that column tells us
#  the buyer name to embed in the saved filename.
#  101 → IHAB,  106 → HALA
# ──────────────────────────────────────────────────────────────
KHOZEMA_EMAIL = "khozema.b@example.com"
KHOZEMA_DIVISION_MAP = {
    "101": "IHAB",
    "106": "HALA",
}

# ──────────────────────────────────────────────────────────────
#  PURCH DEPT EMAILS  <-- Fix 3
#  Added to the TO field of the reply when 'tag. girls' is
#  detected in the email body (case-insensitive).
# ──────────────────────────────────────────────────────────────
PURCH_DEPT_EMAILS = [
    "abeer.g@example.com",
    "Nermeen.k@example.com",
    "Dalia.Ahmed@example.com",
    "zaina.n@example.com",
]

_mix_counter: dict = {}


# ──────────────────────────────────────────────────────────────
#  LOGGING
# ──────────────────────────────────────────────────────────────
logging.basicConfig(
    filename=LOG_FILE,
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)

def log(msg: str):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")
    logging.info(msg)


# ──────────────────────────────────────────────────────────────
#  CHANGE PRICE EXCEL CLEANUP
# ──────────────────────────────────────────────────────────────

def _header_matches_removal(header_text: str) -> bool:
    """Return True if this column header should be removed."""
    h = str(header_text).upper().strip()
    for col in COLUMNS_TO_REMOVE:
        if col.upper() in h:
            return True
    return False


def _is_mandatory_row(row_values: list) -> bool:
    """Return True if this row looks like the 'Mandatory...' descriptor row."""
    for val in row_values:
        if val and "MANDATORY" in str(val).upper():
            return True
    return False


def clean_change_price_excel(filepath: str) -> bool:
    """
    Clean a CHANGE PRICE Excel file in-place:
      1. Remove the first 'Mandatory...' descriptor row (row 2 usually)
      2. Remove unwanted columns (Buyer, Arabic Description, Vendor Number, etc.)
    Returns True on success, False on failure.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        # ── Step 1: Find and delete the 'Mandatory...' descriptor row ──
        # Pass A: normal cell scan (rows 1-5)
        mandatory_row_idx = None
        for row_idx in range(1, 6):
            row_vals = [ws.cell(row=row_idx, column=c).value
                        for c in range(1, ws.max_column + 1)]
            if _is_mandatory_row(row_vals):
                mandatory_row_idx = row_idx
                break

        # Pass B: check merged-cell anchors if Pass A found nothing
        # (openpyxl returns None for merged cells except the top-left anchor,
        #  so a merged "Mandatory" header spanning the whole row is missed by Pass A)
        if mandatory_row_idx is None:
            for merge in ws.merged_cells.ranges:
                anchor_val = ws.cell(row=merge.min_row, column=merge.min_col).value
                if anchor_val and "MANDATORY" in str(anchor_val).upper():
                    if merge.min_row <= 5:
                        mandatory_row_idx = merge.min_row
                        break

        if mandatory_row_idx:
            ws.delete_rows(mandatory_row_idx)
            log(f"   🧹 Removed 'Mandatory' descriptor row (was row {mandatory_row_idx})")

        # ── Step 2: Find columns to delete (scan row 1 for headers) ──
        # Collect column indices to delete (highest-first to avoid index shift)
        cols_to_delete = []
        for col_idx in range(1, ws.max_column + 1):
            header_val = ws.cell(row=1, column=col_idx).value
            if header_val and _header_matches_removal(header_val):
                cols_to_delete.append(col_idx)
                log(f"   🧹 Marking column for removal: '{header_val}' (col {col_idx})")

        # Delete from right to left so indices stay valid
        for col_idx in sorted(cols_to_delete, reverse=True):
            ws.delete_cols(col_idx)

        # ── Step 3: Clean VENDOR ITEM column — blank cells that contain
        #    ONLY meaningless characters (-, comma, @, dot).
        #    Real values like 'AAK-001', 'ABC123', '001' are kept as-is.
        #    Matches all known header variants (with/without S, typos, newlines, etc.) ──
        VENDOR_ITEM_VARIANTS = {
            "VENDOR ITEM", "VENDOR ITEMS", "VENDOR ITEM NUMBER",
            "VENDORITEM", "VENDOR_ITEM", "VENDOR ITEM", "VENDORE ITEM",
            "VENDER ITEM", "VENDOR ITEM NO",
            "SUPPLIER ITEM", "SUPPLIERITEM",
            "ARTICLE #", "ITEM CODE", "CODE",
        }
        JUNK_ONLY = re.compile(r'^[\-,@.]+$')
        for col_idx in range(1, ws.max_column + 1):
            header_val = ws.cell(row=1, column=col_idx).value
            if header_val:
                # Normalise: collapse whitespace/newlines, strip, uppercase
                normalised = re.sub(r'\s+', ' ', str(header_val)).strip().upper()
                if normalised in VENDOR_ITEM_VARIANTS:
                    for row_idx in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if cell.value is not None:
                            stripped = str(cell.value).strip()
                            if JUNK_ONLY.match(stripped):
                                cell.value = None
                    log(f"   🧹 Cleaned junk values in VENDOR ITEM column '{header_val}' (col {col_idx})")
                    break

        wb.save(filepath)
        log(f"   ✅ Change Price file cleaned: {os.path.basename(filepath)}")
        return True

    except Exception as e:
        log(f"   ⚠️  Could not clean Change Price Excel: {e}")
        return False


# ──────────────────────────────────────────────────────────────
#  DESKTOP NOTIFICATION  — stays until you click Dismiss
# ──────────────────────────────────────────────────────────────

def show_notification(sender_label: str, subject: str,
                      filename: str, dest_folder: str,
                      cleaned: bool = False):
    """Non-blocking dark popup; never auto-closes."""
    def _popup():
        root = tk.Tk()
        root.title("📬 New Email Processed")
        root.attributes("-topmost", True)
        root.resizable(False, False)

        root.update_idletasks()
        sw, sh = root.winfo_screenwidth(), root.winfo_screenheight()
        w, h = 460, 230
        # Sit just below the Mail2Nav control panel (340px tall, shifted up)
        mail2nav_h = 340
        mail2nav_y = max(10, (sh - mail2nav_h) // 2 - 60)
        popup_y    = mail2nav_y + mail2nav_h + 6   # 6px gap below Mail2Nav
        root.geometry(f"{w}x{h}+{sw - w - 20}+{popup_y}")

        BG = "#1e1e2e"; ACCENT = "#89b4fa"
        TEXT = "#cdd6f4"; SUBTEXT = "#a6adc8"; GREEN = "#a6e3a1"; YELLOW = "#f9e2af"

        root.configure(bg=BG)
        tk.Frame(root, bg=ACCENT, height=4).pack(fill="x", side="top")

        body = tk.Frame(root, bg=BG, padx=16, pady=12)
        body.pack(fill="both", expand=True)

        tf = tkfont.Font(family="Segoe UI", size=11, weight="bold")
        nf = tkfont.Font(family="Segoe UI", size=9)
        sf = tkfont.Font(family="Segoe UI", size=8)

        tk.Label(body, text="📬  New email processed",
                 font=tf, bg=BG, fg=ACCENT, anchor="w").pack(fill="x")
        tk.Frame(body, bg=SUBTEXT, height=1).pack(fill="x", pady=(4, 8))

        def info_row(label_text, value_text, fg=TEXT):
            f = tk.Frame(body, bg=BG)
            f.pack(fill="x", pady=1)
            tk.Label(f, text=label_text, width=9, anchor="w",
                     font=nf, bg=BG, fg=SUBTEXT).pack(side="left")
            display = value_text[:52] + "…" if len(value_text) > 52 else value_text
            tk.Label(f, text=display, anchor="w",
                     font=nf, bg=BG, fg=fg).pack(side="left")

        info_row("From:",    sender_label)
        info_row("Subject:", subject)
        info_row("Saved:",   filename, fg=GREEN)
        info_row("Folder:",  os.path.basename(dest_folder))

        if cleaned:
            info_row("Cleaned:", "✓ Buyer, Arabic Desc, Vendor No. removed", fg=YELLOW)

        bf = tk.Frame(body, bg=BG)
        bf.pack(fill="x", pady=(10, 0))
        tk.Button(bf, text="Dismiss", command=root.destroy,
                  font=sf, bg=ACCENT, fg=BG, relief="flat",
                  padx=12, pady=4, cursor="hand2").pack(side="right")

        root.mainloop()

    threading.Thread(target=_popup, daemon=True).start()


# ──────────────────────────────────────────────────────────────
#  TEXT KEYWORD SCAN
# ──────────────────────────────────────────────────────────────

def detect_type_from_text(text_sources: list) -> tuple:
    combined = " ".join(text_sources).upper()
    for keyword, req_type, folder in KEYWORD_DETECTION:
        if keyword.upper() in combined:
            return req_type, folder
    return None, None


# ──────────────────────────────────────────────────────────────
#  EXCEL UTILITIES
# ──────────────────────────────────────────────────────────────

def _build_col_map(ws) -> dict:
    mapping = {}
    for row_idx in (1, 2):
        try:
            for cell in ws[row_idx]:
                if cell.value:
                    key = " ".join(str(cell.value).upper().strip().split())
                    mapping[key] = cell.column
        except Exception:
            pass
    return mapping


# All known header variants that mean "Item Number" column
_ITEM_NO_VARIANTS = re.compile(
    r'^('
    r'ITEM\s*NO\.?'           # ITEM NO  ITEM NO.
    r'|ITEM\s*NUM(BER)?'      # ITEM NUM  ITEM NUMBER
    r'|ITEMNO\.?'             # ITEMNO  ITEMNO.
    r'|ITEM\s*#'              # ITEM#
    r'|ITEM\s*\|?\s*V'        # ITEM|V  ITEM V  (Zeina's header)
    r'|ITEM'                  # ITEM  (plain — used by Zaina and similar suppliers)
    r'|NO\.'                  # No.
    r'|SKU'                   # SKU
    r'|STOCK\s*CODE'          # STOCK CODE
    r'|MARJI\s*ITEM\s*(NO\.?)?' # Marji Item No.
    r'|PRODUCT\s*NO\.?'       # PRODUCT NO
    r')$',
    re.IGNORECASE,
)

def _find_item_no_col(ws, max_header_rows: int = 6) -> int | None:
    """
    Scan the first `max_header_rows` rows for any known ITEM NO header variant.
    Returns the 1-based column index, or None if not found.
    """
    for row_idx in range(1, max_header_rows + 1):
        try:
            for cell in ws[row_idx]:
                if cell.value:
                    h = " ".join(str(cell.value).strip().split())  # normalise whitespace
                    if _ITEM_NO_VARIANTS.match(h):
                        return cell.column
        except Exception:
            pass
    return None


def _col_has_data(ws, col_idx: int, start_row: int = 3, max_rows: int = 300) -> bool:
    for row in ws.iter_rows(min_row=start_row, max_row=max_rows,
                             min_col=col_idx, max_col=col_idx,
                             values_only=True):
        val = row[0]
        if val is not None and str(val).strip() not in ("", "-"):
            return True
    return False


def analyze_amro_excel(filepath: str) -> tuple:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        col_map = _build_col_map(ws)

        item_no_col = _find_item_no_col(ws)

        prefixes      = []   # letter prefixes  e.g. ["W", "AF"]
        numeric_codes = []   # numeric prefixes e.g. ["101"]
        if item_no_col:
            for row in ws.iter_rows(min_row=3, max_row=300,
                                    min_col=item_no_col, max_col=item_no_col,
                                    values_only=True):
                val = str(row[0]).strip() if row[0] not in (None, "", "-") else ""
                if not val:
                    continue
                # Letter prefix: W-30044, AF-1234, etc.
                m_alpha = re.match(r'^([A-Za-z]+)', val)
                if m_alpha:
                    p = m_alpha.group(1).upper()
                    if p not in prefixes:
                        prefixes.append(p)
                # Numeric prefix: 101-5500, 202-1234, etc.
                m_num = re.match(r'^(\d+)', val)
                if m_num:
                    n = m_num.group(1)
                    if n not in numeric_codes:
                        numeric_codes.append(n)

        ni_score = cp_score = 0
        for header, col_idx in col_map.items():
            for ni_col in AMRO_NEW_ITEM_COLS:
                if ni_col in header and _col_has_data(ws, col_idx):
                    ni_score += 1
                    break
            for cp_col in AMRO_CHANGE_PRICE_COLS:
                if cp_col in header and _col_has_data(ws, col_idx):
                    cp_score += 1
                    break

        wb.close()
        log(f"   🔬 Amro Excel → NI_score={ni_score} CP_score={cp_score} prefixes={prefixes} numeric_codes={numeric_codes}")

        req_type = "NEW ITEMS" if ni_score > cp_score else "CHANGE PRICE"
        return req_type, prefixes, numeric_codes

    except Exception as e:
        log(f"   ⚠️  analyze_amro_excel error: {e}")
        return None, [], []


def peek_excel_headers(filepath: str) -> list:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=2, values_only=True):
            for cell in row:
                if cell:
                    headers.append(str(cell).upper())
        wb.close()
        return headers
    except Exception as e:
        log(f"   ⚠️  peek_excel_headers error: {e}")
        return []


# ──────────────────────────────────────────────────────────────
#  ITEM PREFIX → FILENAME SEGMENT
# ──────────────────────────────────────────────────────────────

def build_prefix_segment(prefixes: list, label: str) -> str:
    if not prefixes:
        return "MIX"
    if len(prefixes) == 1:
        return prefixes[0]
    today = datetime.now().strftime("%d-%m-%y")
    key = f"{label}-{today}"
    _mix_counter[key] = _mix_counter.get(key, 0) + 1
    return f"MIX{_mix_counter[key]}"


def build_buyer_suffix(numeric_codes: list) -> str:
    """
    Look up each numeric code in BUYER_CODE_MAP.
    Returns a space-separated string of matched buyer names, or "" if none.
    Example: ["101"] → "IHAB"
             ["101", "202"] → "IHAB SARA"
    """
    names = []
    for code in numeric_codes:
        name = BUYER_CODE_MAP.get(code)
        if name and name not in names:
            names.append(name)
    return " ".join(names)


# ──────────────────────────────────────────────────────────────
#  FILE SAVE HELPERS
# ──────────────────────────────────────────────────────────────

def move_file_unique(src: str, dest_folder: str, save_name: str) -> str:
    os.makedirs(dest_folder, exist_ok=True)
    dest = os.path.join(dest_folder, save_name)
    if os.path.exists(dest):
        base, ext = os.path.splitext(save_name)
        c = 1
        while os.path.exists(dest):
            dest = os.path.join(dest_folder, f"{base}_{c}{ext}")
            c += 1
    shutil.move(src, dest)
    return dest


def save_attachment_direct(att, dest_folder: str, save_name: str) -> str:
    os.makedirs(dest_folder, exist_ok=True)
    dest = os.path.join(dest_folder, save_name)
    if os.path.exists(dest):
        base, ext = os.path.splitext(save_name)
        c = 1
        while os.path.exists(dest):
            dest = os.path.join(dest_folder, f"{base}_{c}{ext}")
            c += 1
    att.SaveAsFile(dest)
    return dest


def _tmp_path(name: str) -> str:
    return os.path.join(os.environ.get("TEMP", r"C:\Temp"), f"__atch_tmp_{name}")


# ──────────────────────────────────────────────────────────────
#  POST-SAVE CLEANUP: apply to CHANGE PRICE files only
# ──────────────────────────────────────────────────────────────

def maybe_clean_change_price(filepath: str, req_type: str) -> bool:
    """If this is a CHANGE PRICE Excel file, clean it and return True if cleaned."""
    if req_type != "CHANGE PRICE":
        return False
    ext = os.path.splitext(filepath)[1].lower()
    if ext not in ('.xlsx', '.xls'):
        return False
    return clean_change_price_excel(filepath)


# ──────────────────────────────────────────────────────────────
#  NEW ITEMS EXCEL CLEANUP
#  Strips junk/template rows that have no ITEM NO value.
#  Amro's template pre-fills BARCODE and other columns for ~40
#  rows even when only 1 real item exists — those ghost rows
#  must be removed so Navision only sees real data.
# ──────────────────────────────────────────────────────────────

def clean_new_items_excel(filepath: str) -> bool:
    """
    Clean a NEW ITEMS Excel file in-place:
      1. Remove the Mandatory header row if present
      2. Delete every row whose ITEM NO cell is blank after the header
    Returns True on success, False on failure.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        # Step 1: Remove Mandatory row (merged or plain)
        mandatory_row_idx = None
        for row_idx in range(1, 6):
            row_vals = [ws.cell(row=row_idx, column=c).value
                        for c in range(1, ws.max_column + 1)]
            if _is_mandatory_row(row_vals):
                mandatory_row_idx = row_idx
                break
        if mandatory_row_idx is None:
            for merge in ws.merged_cells.ranges:
                anchor_val = ws.cell(row=merge.min_row, column=merge.min_col).value
                if anchor_val and "MANDATORY" in str(anchor_val).upper():
                    if merge.min_row <= 5:
                        mandatory_row_idx = merge.min_row
                        break
        if mandatory_row_idx:
            ws.delete_rows(mandatory_row_idx)
            log(f"   Removed Mandatory row (was row {mandatory_row_idx})")

        else:
            # ── No Mandatory row — but header row may still have merged cells ──
            # (e.g. "BARCODE" spanning multiple columns in Amro's NEW ITEMS template)
            merges_in_header_ni = [
                m for m in list(ws.merged_cells.ranges)
                if m.min_row <= 1 <= m.max_row
            ]
            for merge in merges_in_header_ni:
                anchor_val = ws.cell(row=merge.min_row, column=merge.min_col).value
                ws.unmerge_cells(str(merge))
                ws.cell(row=merge.min_row, column=merge.min_col).value = anchor_val
            if merges_in_header_ni:
                log(f"   🔓 [NI] Unmerged {len(merges_in_header_ni)} header merge(s) — each column now has its own header")

        # Step 2: Find ITEM NO column in the header row
        item_no_col = None
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col_idx).value
            if val and _ITEM_NO_VARIANTS.match(str(val).strip().upper()):
                item_no_col = col_idx
                break

        if not item_no_col:
            log("   [NI] ITEM NO column not found - skipping empty-row cleanup")
            wb.save(filepath)
            return True

        # Step 3: Delete rows with no ITEM NO (bottom-up so indices stay valid)
        removed = 0
        for row_idx in range(ws.max_row, 1, -1):
            cell_val = ws.cell(row=row_idx, column=item_no_col).value
            if cell_val is None or str(cell_val).strip() in ("", "-", "None"):
                ws.delete_rows(row_idx)
                removed += 1

        log(f"   Removed {removed} empty/junk row(s) - {ws.max_row - 1} real item(s) remain")
        wb.save(filepath)
        log(f"   New Items file cleaned: {os.path.basename(filepath)}")
        return True

    except Exception as e:
        log(f"   clean_new_items_excel error: {e}")
        return False


def maybe_clean_new_items(filepath: str, req_type: str) -> bool:
    """If this is a NEW ITEMS Excel file, strip empty rows and return True if cleaned."""
    if req_type != "NEW ITEMS":
        return False
    ext = os.path.splitext(filepath)[1].lower()
    if ext not in ('.xlsx', '.xls'):
        return False
    return clean_new_items_excel(filepath)


# ──────────────────────────────────────────────────────────────
#  NEW ITEMS EXCEL CLEANUP
#  Strips junk/template rows that have no ITEM NO value.
#  Amro's template pre-fills BARCODE and other columns for ~40
#  rows even when only 1 real item exists — those ghost rows
#  must be removed so Navision only sees real data.
# ──────────────────────────────────────────────────────────────

def clean_new_items_excel(filepath: str) -> bool:
    """
    Clean a NEW ITEMS Excel file in-place:
      1. Remove the Mandatory header row if present
      2. Delete every row whose ITEM NO cell is blank after the header
    Returns True on success, False on failure.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        # Step 1: Remove Mandatory row (merged or plain)
        mandatory_row_idx = None
        for row_idx in range(1, 6):
            row_vals = [ws.cell(row=row_idx, column=c).value
                        for c in range(1, ws.max_column + 1)]
            if _is_mandatory_row(row_vals):
                mandatory_row_idx = row_idx
                break
        if mandatory_row_idx is None:
            for merge in ws.merged_cells.ranges:
                anchor_val = ws.cell(row=merge.min_row, column=merge.min_col).value
                if anchor_val and "MANDATORY" in str(anchor_val).upper():
                    if merge.min_row <= 5:
                        mandatory_row_idx = merge.min_row
                        break
        if mandatory_row_idx:
            ws.delete_rows(mandatory_row_idx)
            log(f"   🧹 [NI] Removed Mandatory row (was row {mandatory_row_idx})")

            # ── Remove ALL remaining merged cells after Mandatory row deletion ──
            # openpyxl shifts merge definitions when a row is deleted, causing
            # header merges (ITEM NO, BARCODE etc.) to overlap the data rows.
            # Save header values first, unmerge all, then restore.
            header_values_ni = {}
            for c in range(1, ws.max_column + 1):
                header_values_ni[c] = ws.cell(row=1, column=c).value

            merges_to_remove_ni = list(ws.merged_cells.ranges)
            for merge in merges_to_remove_ni:
                ws.unmerge_cells(str(merge))

            for c, val in header_values_ni.items():
                ws.cell(row=1, column=c).value = val
            log(f"   🔓 [NI] Removed {len(merges_to_remove_ni)} merged cell range(s), headers restored")

        else:
            # ── No Mandatory row — but header row may still have merged cells ──
            # (e.g. "BARCODE" spanning multiple columns in Amro's NEW ITEMS template)
            merges_in_header_ni = [
                m for m in list(ws.merged_cells.ranges)
                if m.min_row <= 1 <= m.max_row
            ]
            for merge in merges_in_header_ni:
                anchor_val = ws.cell(row=merge.min_row, column=merge.min_col).value
                ws.unmerge_cells(str(merge))
                ws.cell(row=merge.min_row, column=merge.min_col).value = anchor_val
            if merges_in_header_ni:
                log(f"   🔓 [NI] Unmerged {len(merges_in_header_ni)} header merge(s) — each column now has its own header")

        # Step 2: Find ITEM NO column in the header row
        item_no_col = None
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col_idx).value
            if val and _ITEM_NO_VARIANTS.match(str(val).strip().upper()):
                item_no_col = col_idx
                break

        if not item_no_col:
            log("   ⚠️  [NI] ITEM NO column not found — skipping empty-row cleanup")
            wb.save(filepath)
            return True

        # Step 3: Delete rows with no ITEM NO (bottom-up so indices stay valid)
        removed = 0
        for row_idx in range(ws.max_row, 1, -1):
            cell_val = ws.cell(row=row_idx, column=item_no_col).value
            if cell_val is None or str(cell_val).strip() in ("", "-", "None"):
                ws.delete_rows(row_idx)
                removed += 1

        log(f"   🧹 [NI] Removed {removed} empty/junk row(s) — {ws.max_row - 1} real item(s) remain")
        wb.save(filepath)
        log(f"   ✅ New Items file cleaned: {os.path.basename(filepath)}")
        return True

    except Exception as e:
        log(f"   ⚠️  clean_new_items_excel error: {e}")
        return False


def maybe_clean_new_items(filepath: str, req_type: str) -> bool:
    """If this is a NEW ITEMS Excel file, strip empty rows and return True if cleaned."""
    if req_type != "NEW ITEMS":
        return False
    ext = os.path.splitext(filepath)[1].lower()
    if ext not in ('.xlsx', '.xls'):
        return False
    return clean_new_items_excel(filepath)


# ──────────────────────────────────────────────────────────────
#  REVIEW EXCEL — BARCODE HIGHLIGHT
#  Scans the Operations folder for the REVIEW_NEW_ITEMS file
#  and highlights the BARCODE column in pink for easy reading.
# ──────────────────────────────────────────────────────────────

def highlight_review_barcodes(operations_folder: str) -> bool:
    """
    Find the REVIEW_NEW_ITEMS Excel file in operations_folder and
    highlight all cells in the BARCODE column (header + data) in pink.
    Returns True if successfully highlighted, False otherwise.
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill

        # Find the review file (matches REVIEW_NEW_ITEMS*.xlsx)
        review_file = None
        for fname in os.listdir(operations_folder):
            if re.match(r'REVIEW_NEW_ITEMS', fname, re.IGNORECASE) and \
               fname.lower().endswith(('.xlsx', '.xls')):
                review_file = os.path.join(operations_folder, fname)
                break

        if not review_file:
            log("   INFO: No REVIEW_NEW_ITEMS file found — skipping barcode highlight.")
            return False

        wb = openpyxl.load_workbook(review_file)
        ws = wb.active

        # Find BARCODE column header (row 1)
        barcode_col = None
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col_idx).value
            if val and "BARCODE" in str(val).upper():
                barcode_col = col_idx
                break

        if not barcode_col:
            log("   WARNING: BARCODE column not found in Review file.")
            wb.close()
            return False

        # Pink fill matching the screenshot style
        pink_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")

        highlighted = 0
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=barcode_col)
            cell.fill = pink_fill
            highlighted += 1

        wb.save(review_file)
        log(f"   OK Highlighted {highlighted} barcode cell(s) in '{os.path.basename(review_file)}'")
        return True

    except Exception as e:
        log(f"   WARNING: highlight_review_barcodes error: {e}")
        return False


# ──────────────────────────────────────────────────────────────
#  ARCHIVE MOVER
#  After processing, move files from Buyers sub-folder to
#  Input Archive\<date> folder.
# ──────────────────────────────────────────────────────────────

def archive_buyers_files(operations_folder: str):
    r"""
    Move all files from Operations\Buyers  ->  Operations\Input Archive\DD-MM-YY
    Creates the date folder if it doesn't exist.
    """
    buyers_folder  = os.path.join(operations_folder, "Buyers")
    today_str      = datetime.now().strftime("%d-%m-%y")
    archive_folder = os.path.join(operations_folder, "Input Archive", today_str)

    if not os.path.isdir(buyers_folder):
        log(f"   INFO: Buyers folder not found — skipping archive: {buyers_folder}")
        return

    files = [f for f in os.listdir(buyers_folder)
             if os.path.isfile(os.path.join(buyers_folder, f))]

    if not files:
        log("   INFO: Buyers folder is empty — nothing to archive.")
        return

    os.makedirs(archive_folder, exist_ok=True)
    moved = 0
    for fname in files:
        src  = os.path.join(buyers_folder, fname)
        dest = os.path.join(archive_folder, fname)
        # Avoid overwrite — add counter suffix if needed
        if os.path.exists(dest):
            base, ext = os.path.splitext(fname)
            c = 1
            while os.path.exists(dest):
                dest = os.path.join(archive_folder, f"{base}_{c}{ext}")
                c += 1
        shutil.move(src, dest)
        log(f"   ARCHIVED: '{fname}' -> Input Archive\\{today_str}")
        moved += 1

    log(f"OK Archive done — {moved} file(s) moved to '{archive_folder}'")


# ──────────────────────────────────────────────────────────────
#  BLOCK PURCHASE HANDLER
#  Reads "Marji Item No." column from Excel → writes item codes
#  to BLOCK_TXT_FILE (one per line, no header, no extra data).
# ──────────────────────────────────────────────────────────────

def _is_block_purchase_subject(subject: str) -> bool:
    """
    Return True if the email subject indicates a Block Purchase request.
    Handles case variations: "Block Purchase", "block purchase", "BLOCK PURCHASE".
    """
    return bool(re.search(r'block\s*purchase', subject, re.IGNORECASE))


def extract_marji_items_from_excel(filepath: str) -> list:
    """
    Open an Excel file and extract all values from the "Marji Item No." column.

    Rules:
      - Search rows 1–6 for the header (exact name "Marji Item No." matched
        case-insensitively; also accepts "MARJI ITEM NO", "MARJI ITEM NUMBER")
      - Copy every non-empty cell below that header until 10 consecutive
        blank cells are found (end-of-data guard)
      - Return list of strings like ["MS2391", "MS2476", ...]
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active

        marji_col  = _find_item_no_col(ws, max_header_rows=6)
        header_row = None

        # Find which row contains the header
        if marji_col:
            for row_idx in range(1, 7):
                if ws.cell(row=row_idx, column=marji_col).value:
                    header_row = row_idx
                    log(f"   📋 Marji Item No. found at row={row_idx} col={marji_col}")
                    break

        if not marji_col:
            log("   ⚠️  'Marji Item No.' column not found in Excel.")
            wb.close()
            return []

        items       = []
        blank_count = 0

        for row in ws.iter_rows(min_row=header_row + 1, max_row=2000,
                                 min_col=marji_col, max_col=marji_col,
                                 values_only=True):
            val = row[0]
            if val is None or str(val).strip() == "":
                blank_count += 1
                if blank_count >= 10:
                    break           # 10 consecutive blanks = end of data
                continue
            blank_count = 0
            items.append(str(val).strip())

        wb.close()
        log(f"   ✅ Extracted {len(items)} Marji item(s) from Excel.")
        return items

    except Exception as e:
        log(f"   ⚠️  extract_marji_items_from_excel error: {e}")
        return []


def handle_block_purchase(att, sender_email: str, subject: str) -> bool:
    """
    Handle a Block Purchase email:
      1. Save Excel attachment to temp
      2. Extract Marji Item No. values
      3. Append (or create) BLOCK_TXT_FILE with one item code per line
      4. Show notification
    Returns True if handled, False if skipped.
    """
    original_name = att.FileName
    ext = os.path.splitext(original_name)[1].lower()
    if ext not in ('.xlsx', '.xls'):
        log(f"   ⏭️  Block Purchase: skipping non-Excel file '{original_name}'")
        return False

    tmp = _tmp_path(original_name)
    try:
        att.SaveAsFile(tmp)
    except Exception as e:
        log(f"   ⚠️  Block Purchase: could not save temp: {e}")
        return False

    items = extract_marji_items_from_excel(tmp)
    try:
        os.remove(tmp)
    except Exception:
        pass

    if not items:
        log(f"   ⚠️  Block Purchase: no items found in '{original_name}' — file not written.")
        show_notification(
            sender_email, subject,
            "⚠️ No items found — check 'Marji Item No.' column",
            os.path.dirname(BLOCK_TXT_FILE),
        )
        return True

    # Write to txt — APPEND so multiple emails accumulate, separated by a blank line
    try:
        os.makedirs(os.path.dirname(BLOCK_TXT_FILE), exist_ok=True)
        already_exists = os.path.isfile(BLOCK_TXT_FILE)
        with open(BLOCK_TXT_FILE, "a", encoding="utf-8") as f:
            if already_exists:
                f.write("\n")   # blank separator between batches
            for item in items:
                f.write(item + "\n")
        log(
            f"✅ BLOCK   | From: {sender_email} | Subject: {subject}\n"
            f"           | Items written: {len(items)}\n"
            f"           | File: {BLOCK_TXT_FILE}"
        )
    except Exception as e:
        log(f"   ❌ Block Purchase: could not write txt: {e}")
        show_notification(sender_email, subject, f"❌ Write error: {e}", "")
        return True

    show_notification(
        sender_email,
        subject,
        f"✓ {len(items)} item(s) → {os.path.basename(BLOCK_TXT_FILE)}",
        os.path.dirname(BLOCK_TXT_FILE),
        cleaned=False,
    )
    return True


# ──────────────────────────────────────────────────────────────
#  KHOZEMA HANDLER  (Fix 1, 2, 3, 4, 5)
# ──────────────────────────────────────────────────────────────

def extract_division_from_excel(filepath: str, ext: str) -> str:
    """
    Fix 1 & 5: Read the DIV or DIVISION column from khozema's Excel file.
    Returns the division code string (e.g. '101' or '106'), or '' if not found.
    """
    if ext not in ('.xlsx', '.xls'):
        return ""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        div_col    = None
        header_row = None
        for r in range(1, 6):
            for cell in ws[r]:
                if cell.value and re.match(r'^DIV(ISION)?$',
                                           str(cell.value).strip(), re.IGNORECASE):
                    div_col    = cell.column
                    header_row = r
                    break
            if div_col:
                break
        if div_col is None:
            wb.close()
            return ""
        for row in ws.iter_rows(min_row=header_row + 1, max_row=500,
                                 min_col=div_col, max_col=div_col,
                                 values_only=True):
            val = row[0]
            if val is not None and str(val).strip():
                wb.close()
                return str(val).strip()
        wb.close()
        return ""
    except Exception as e:
        log(f"   ⚠️  extract_division_from_excel error: {e}")
        return ""


def _detect_tag_girls(body: str) -> bool:
    """Return True if body contains 'tag. girls' (case-insensitive)."""
    return bool(re.search(r'tag\.\s*girls', body, re.IGNORECASE))


def handle_khozema(att, message, sender_email: str, subject: str,
                   body: str, entry_id: str = "") -> bool:
    """
    Handler for khozema.b@example.com.

    Fix 1 & 5: Read DIV/DIVISION column from Excel → map to buyer name → embed in filename.
    Fix 2:     Reply All (keep all original To + CC recipients).
    Fix 3:     'tag. girls' in body → add PURCH_DEPT_EMAILS to TO of reply.
    Fix 4:     'tag. girls' → add PO line after completion sentence in reply body.
    """
    original_name = att.FileName
    ext = os.path.splitext(original_name)[1].lower()
    if ext not in ('.xlsx', '.xls', '.csv', '.txt'):
        return False

    today = datetime.now().strftime("%d-%m-%y")

    # Detect request type from subject / body / filename
    req_type, dest_folder = detect_type_from_text([subject, body, original_name])
    if req_type is None:
        req_type    = "NEW ITEMS"
        dest_folder = OPERATIONS_FOLDER

    # Save to temp so we can read inside the Excel
    tmp = _tmp_path(original_name)
    try:
        att.SaveAsFile(tmp)
    except Exception as e:
        log(f"   ⚠️  Khozema: could not save temp: {e}")
        return False

    # Fix 1 & 5: get buyer name from DIV/DIVISION column
    div_code   = extract_division_from_excel(tmp, ext)
    buyer_name = KHOZEMA_DIVISION_MAP.get(str(div_code).strip(), "")
    if not buyer_name:
        # Fallback: look for code in subject / filename
        buyer_name = extract_buyer_from_text(f"{subject} {original_name}".upper())
    log(f"   📋 Khozema div_code='{div_code}' buyer='{buyer_name}'")

    # Build filename: e.g. "NEW ITEMS MAYB HALA 13-04-26.xlsx"
    prefixes = extract_letter_prefix_from_excel(tmp, ext)
    parts    = [req_type]
    if prefixes:
        parts.append(build_prefix_segment(prefixes, "KHOZEMA"))
    if buyer_name:
        parts.append(buyer_name)
    parts.append(today)
    new_name = " ".join(parts) + ext

    try:
        final = move_file_unique(tmp, dest_folder, new_name)
    except Exception as e:
        log(f"   ⚠️  Khozema: move failed ({e}); re-saving directly.")
        try:
            att.SaveAsFile(os.path.join(dest_folder, new_name))
            final = os.path.join(dest_folder, new_name)
        except Exception as e2:
            log(f"   ❌ Khozema: could not save: {e2}")
            return True

    cleaned = maybe_clean_change_price(final, req_type)

    log(
        f"✅ KHOZEMA | Subject: {subject}\n"
        f"           | Original: '{original_name}'\n"
        f"           | Type: '{req_type}'\n"
        f"           | Buyer: '{buyer_name}'\n"
        f"           | Saved as: '{os.path.basename(final)}'\n"
        f"           | Folder: {dest_folder}"
        + (f"\n           | Cleaned: Yes" if cleaned else "")
    )
    show_notification(f"Khozema ({sender_email})", subject,
                      os.path.basename(final), dest_folder, cleaned=cleaned)
    _register_job(entry_id, sender_email, subject, os.path.basename(final), req_type)

    # ── Fix 2, 3, 4: build and send reply ─────────────────────
    tag_girls = _detect_tag_girls(body)
    try:
        import pythoncom as _pc
        _pc.CoInitialize()
        _outlook  = win32com.client.Dispatch("Outlook.Application")
        _ns       = _outlook.GetNamespace("MAPI")
        _original = _ns.GetItemFromID(entry_id)

        # Fix 2: ReplyAll keeps every original To + CC recipient
        reply = _original.ReplyAll()

        # Fix 3: add Purch Dept. to TO when 'tag. girls' detected
        if tag_girls:
            extra = "; ".join(PURCH_DEPT_EMAILS)
            reply.To = ((reply.To or "") + "; " + extra).strip("; ")
            log("   📧 'tag. girls' detected — Purch Dept. added to TO")

        # Fix 4: build body with optional PO line
        first_name = "Khozema"
        po_line    = (
            "<br><br>@Purchasing Team, kindly proceed with issuing the purchase orders."
            if tag_girls else ""
        )
        completion = (
            "Noted, I will review and process this accordingly."
            if req_type == "CHANGE PRICE"
            else "The new items creation has been completed."
        )
        existing_html = reply.HTMLBody or ""
        reply.HTMLBody = (
            f"<div style='font-family:Calibri,sans-serif;font-size:11pt'>"
            f"Dear {first_name},<br><br>"
            f"{completion}"
            f"{po_line}"
            f"<br><br>Thank you."
            f"<br><br>Best regards,<br>Abdul Khader"
            f"</div><br>"
        ) + existing_html

        reply.Display(True)
        log(f"   ✅ Khozema reply opened in Outlook")
        _pc.CoUninitialize()
    except Exception as e:
        log(f"   ⚠️  Khozema reply error: {e}")

    return True


# ──────────────────────────────────────────────────────────────
#  AMRO HANDLER
# ──────────────────────────────────────────────────────────────

def clean_amro_change_price_columns(filepath: str):
    """
    For Amro's CHANGE PRICE files — full cleanup in 5 steps.

    Amro's actual file structure (confirmed from real file inspection):
      Row 1  = Mandatory row — 4 merged cells spanning cols 1-2, 3-15, 16-17, 18-20
               Text: "Mandatory", "New Item", "Changing Price", "Changing Pur. Cost"
      Row 2  = Real headers:
               A=ITEM NO, B=IMAGES, C=BARCODE, D=VENDOR ITEMS, E=ENGLISH DESC,
               F=Base UOM, G=Cost Price, H=Selling Price (hidden), I=Item Category (hidden),
               J=Sub Category (hidden), K=Sub-Sub Category (hidden), L=BUYER, M=Arabic Desc,
               N=Vender Number, O=Line Disc%, P=OLD RSP, Q=NEW RSP, R=Old Cost,
               S=New Cost, T=Line Disc, U onwards = lookup/category columns (delete all)
      Row 3+ = Data

    Required output — exactly 7 columns in this order:
      A=ITEM NO | B=IMAGES | C=OLD RSP | D=NEW RSP | E=Old Cost | F=New Cost | G=Line Disc

    Steps:
      1. Remove Mandatory row (row 1) — merged or plain, safely skipped if absent
      2. Now row 2 becomes row 1 (real headers). Delete ALL columns from col 21 (U) onward.
      3. Delete cols C through O (positions 3-15) — BARCODE, VENDOR ITEMS, DESC, hidden cols,
         BUYER, Arabic, Vendor No, Line Disc% — keeping only A, B, P→T
         After deletion: A=ITEM NO, B=IMAGES, C=OLD RSP, D=NEW RSP, E=Old Cost,
                         F=New Cost, G=Line Disc
      4. Auto-write "IMAGES" header in col B if it is empty
      5. Clear any text values from col B data cells (keep embedded images intact)
    """
    # ── Canonical header names to write in the final output ──────────────────
    # These rename whatever Amro used to a clean, consistent label.
    CANONICAL_NAMES = {
        "OLD RSP":    {"OLD RSP"},
        "NEW RSP":    {"NEW RSP"},
        "OLD COST":   {"OLD COST", "OLD COST ", "OLDCOST"},
        "NEW COST":   {"NEW COST", "NEW COST ", "NEWCOST"},
        "LINE DISC":  {"LINE DISC", "LINE DISC ", "LINE DISC %", "LINEDISC"},
    }

    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        # ── Step 1: Remove Mandatory descriptor row ───────────────────────────
        # Pass A: merged-cell anchor — Amro always uses merged cells for row 1
        mandatory_row_idx = None
        for merge in ws.merged_cells.ranges:
            anchor_val = ws.cell(row=merge.min_row, column=merge.min_col).value
            if anchor_val and "MANDATORY" in str(anchor_val).upper():
                if merge.min_row <= 5:
                    mandatory_row_idx = merge.min_row
                    log(f"   🧹 Amro CP: found Mandatory merged row at row {mandatory_row_idx}")
                    break

        # Pass B: fallback — plain (non-merged) scan rows 1-5
        if mandatory_row_idx is None:
            for row_idx in range(1, 6):
                row_vals = [ws.cell(row=row_idx, column=c).value
                            for c in range(1, min(ws.max_column + 1, 10))]
                if _is_mandatory_row(row_vals):
                    mandatory_row_idx = row_idx
                    log(f"   🧹 Amro CP: found Mandatory plain row at row {mandatory_row_idx}")
                    break

        if mandatory_row_idx:
            ws.delete_rows(mandatory_row_idx)
            log(f"   🧹 Amro CP: deleted Mandatory row (was row {mandatory_row_idx})")

            # ── Fix A: Remove ALL remaining merged cells ──────────────────────
            # When we delete row 1, openpyxl shifts the 4 merge definitions
            # so they now overlap the header row — causing ITEM NO and OLD RSP
            # to appear merged. Unmerging resets those cells to None, so we
            # must save all header values FIRST, unmerge, then restore them.
            header_values = {}
            for c in range(1, ws.max_column + 1):
                header_values[c] = ws.cell(row=1, column=c).value

            merges_to_remove = list(ws.merged_cells.ranges)
            for merge in merges_to_remove:
                ws.unmerge_cells(str(merge))

            # Restore header values that were wiped by unmerge
            for c, val in header_values.items():
                ws.cell(row=1, column=c).value = val
            log(f"   🔓 Amro CP: removed {len(merges_to_remove)} merged cell range(s), headers restored")

            # ── Fix B: Shift image row anchors by -1 ─────────────────────────
            # openpyxl does NOT automatically update image anchors when rows are
            # deleted. Every image anchor's row is 0-based, so we subtract 1
            # from both _from.row and to.row to keep images in the correct row.
            shifted = 0
            for img in ws._images:
                try:
                    a = img.anchor
                    if a._from.row > 0:
                        a._from.row -= 1
                        shifted += 1
                    if a.to.row > 0:
                        a.to.row -= 1
                except Exception as e:
                    log(f"   ⚠️  Amro CP: could not shift image anchor: {e}")
            log(f"   🖼️  Amro CP: shifted {shifted} image anchor(s) by -1 row")

        else:
            log("   ℹ️  Amro CP: no Mandatory row found — Amro already removed it")

        # After deletion, row 1 is now the real header row (ITEM NO, IMAGES, ...)
        header_row = 1

        # ── Step 2: Delete ALL columns from position 21 (U) onward ───────────
        # These are Amro's lookup/category columns (101 Houseware, 102 Furniture, etc.)
        # We always delete them — no exceptions.
        total_cols = ws.max_column
        if total_cols > 20:
            for col_idx in range(total_cols, 20, -1):
                ws.delete_cols(col_idx)
            log(f"   ✂️  Amro CP: deleted cols 21→{total_cols} (all lookup/category columns)")
        log(f"   📋 Amro CP: {ws.max_column} column(s) remain after col-T truncation")

        # ── Step 3: Delete cols C through O (positions 3-15) ─────────────────
        # These are: BARCODE, VENDOR ITEMS, ENGLISH DESC, Base UOM, Cost Price,
        #            Selling Price (hidden), Item Category (hidden),
        #            Sub Category (hidden), Sub-Sub Category (hidden),
        #            BUYER, Arabic Desc, Vender Number, Line Disc%
        # We delete positions 3-15 right-to-left so indices stay valid.
        # After this, what was P(16), Q(17), R(18), S(19), T(20)
        # becomes C(3), D(4), E(5), F(6), G(7).
        for col_idx in range(15, 2, -1):  # positions 15 down to 3
            header_val = str(ws.cell(row=header_row, column=col_idx).value or "").strip()
            ws.delete_cols(col_idx)
            log(f"   🗑️  Amro CP: deleted col {col_idx} ('{header_val}')")
        log(f"   📋 Amro CP: {ws.max_column} column(s) remain after middle-col deletion")

        # Now layout is:
        #   Col 1 (A) = ITEM NO
        #   Col 2 (B) = IMAGES
        #   Col 3 (C) = OLD RSP
        #   Col 4 (D) = NEW RSP
        #   Col 5 (E) = Old Cost
        #   Col 6 (F) = New Cost
        #   Col 7 (G) = Line Disc

        # ── Step 4: Auto-write "IMAGES" header in col B if empty ─────────────
        col_b_header = str(ws.cell(row=header_row, column=2).value or "").strip()
        if not col_b_header:
            ws.cell(row=header_row, column=2).value = "IMAGES"
            log("   🖼️  Amro CP: col B header was empty — written 'IMAGES'")
        else:
            log(f"   🖼️  Amro CP: col B header already has value: '{col_b_header}'")

        # ── Step 5: Clear text values from col B data cells ───────────────────
        # Embedded images are drawing objects, NOT cell values — they are unaffected.
        # Only text/number cell values are cleared.
        cleared_text = 0
        for row_idx in range(header_row + 1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=2)
            if cell.value is not None and str(cell.value).strip() != "":
                cell.value = None
                cleared_text += 1
        if cleared_text:
            log(f"   🧹 Amro CP: cleared {cleared_text} text value(s) from IMAGES column (col B)")

        # ── Step 6: Remove all spaces from ITEM NO column (col A) data cells ───
        # e.g. "ALAM 481" → "ALAM481"  /  "W 30044" → "W30044"
        # Applies to every data row — header cell is left untouched.
        fixed_item_no = 0
        for row_idx in range(header_row + 1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=1)
            if cell.value is not None:
                original = str(cell.value)
                cleaned  = original.replace(" ", "")
                if cleaned != original:
                    cell.value = cleaned
                    fixed_item_no += 1
        if fixed_item_no:
            log(f"   🔧 Amro CP: removed spaces from {fixed_item_no} ITEM NO value(s)")

        # ── Step 7: Unmerge ALL remaining merged cells in the header row ─────
        # Amro's real header row (row 2 in original, now row 1) has merged cells
        # like "BARCODE" spanning C-O, "OLD RSP" spanning P-Q, "OLD Cost" spanning R-S.
        # These must be unmerged so each column shows its correct header individually.
        header_values_before = {}
        for c in range(1, ws.max_column + 1):
            header_values_before[c] = ws.cell(row=header_row, column=c).value

        merges_in_header = [
            m for m in list(ws.merged_cells.ranges)
            if m.min_row <= header_row <= m.max_row
        ]
        for merge in merges_in_header:
            anchor_val = ws.cell(row=merge.min_row, column=merge.min_col).value
            ws.unmerge_cells(str(merge))
            # Restore anchor value (unmerge wipes non-anchor cells — anchor is fine)
            ws.cell(row=merge.min_row, column=merge.min_col).value = anchor_val

        if merges_in_header:
            log(f"   🔓 Amro CP: unmerged {len(merges_in_header)} header merge(s) — each column now has its own header")

        # ── Final log: confirm output column layout ───────────────────────────
        final_headers = [
            str(ws.cell(row=header_row, column=c).value or "").strip()
            for c in range(1, ws.max_column + 1)
        ]
        log(f"   ✅ Amro CP: final columns ({ws.max_column}): {final_headers}")

        wb.save(filepath)

    except Exception as e:
        log(f"   ⚠️  clean_amro_change_price_columns error: {e}")


def handle_amro(att, sender_email: str, subject: str,
                body: str, label: str, entry_id: str = "") -> bool:
    original_name = att.FileName
    ext = os.path.splitext(original_name)[1].lower()
    if ext not in ('.xlsx', '.xls', '.csv', '.txt'):
        return False

    today = datetime.now().strftime("%d-%m-%y")

    req_type, dest_folder = detect_type_from_text([subject, body])

    tmp = _tmp_path(original_name)
    try:
        att.SaveAsFile(tmp)
    except Exception as e:
        log(f"   ⚠️  Could not save temp for Amro: {e}")
        return False

    excel_type, prefixes, numeric_codes = analyze_amro_excel(tmp)

    if req_type is None:
        req_type    = excel_type or "CHANGE PRICE"
        dest_folder = (CHANGE_PRICE_FOLDER
                       if req_type == "CHANGE PRICE"
                       else OPERATIONS_FOLDER)

    # Use unified filename builder:
    #   letter prefix  → from Excel ITEM NO column
    #   buyer name     → from subject or original filename via BUYER_CODE_MAP
    new_name = build_clean_filename(req_type, label, tmp, ext,
                                    original_name=original_name,
                                    subject=subject)

    try:
        final = move_file_unique(tmp, dest_folder, new_name)
    except Exception as e:
        log(f"   ⚠️  Move failed ({e}); re-saving directly.")
        try:
            att.SaveAsFile(os.path.join(dest_folder, new_name))
            final = os.path.join(dest_folder, new_name)
        except Exception as e2:
            log(f"   ❌ Could not save: {e2}")
            return True

    # ── Clean CHANGE PRICE file ──
    # ── AHMAD ──
    if "ahmad" in sender_email.lower():
        cleaned = maybe_clean_change_price(final, req_type)

        log(
            f"✅ AHMAD   | Subject: {subject}\n"
            f"           | Original: '{original_name}'\n"
            f"           | Type: '{req_type}'\n"
            f"           | Saved as: '{os.path.basename(final)}'\n"
            f"           | Folder: {dest_folder}"
            + (f"\n           | Cleaned: Yes" if cleaned else "")
        )

        show_notification(
            f"{label} ({sender_email})",
            subject,
            os.path.basename(final),
            dest_folder,
            cleaned=cleaned
        )

        _register_job(entry_id, sender_email, subject, os.path.basename(final), req_type)

        return True

    # ── AMRO ──
    if "amro" in sender_email.lower():
        # Step 1: CHANGE PRICE — Mandatory row removal + whitelist column cleanup.
        #         clean_amro_change_price_columns() handles everything:
        #           a) Removing the merged Mandatory descriptor row (if present)
        #           b) Keeping only known columns: ITEM NO, OLD RSP, NEW RSP,
        #              OLD COST, NEW COST, DISC, BARCODE, IMAGES/PHOTO/PIC
        #           c) Deleting all other junk/blank columns
        #         NOTE: maybe_clean_change_price() is intentionally NOT called for
        #         Amro CHANGE PRICE — clean_amro_change_price_columns handles it all.
        if req_type == "CHANGE PRICE" and os.path.splitext(final)[1].lower() in ('.xlsx', '.xls'):
            clean_amro_change_price_columns(final)
            combined_check = f"{original_name} {subject}".upper()
            if re.search(r'\bCOST\b', combined_check):
                show_change_cost_reminder(sender_name=label)
        # Step 3: NEW ITEMS — strip Mandatory row + ghost rows with no ITEM NO
        elif req_type == "NEW ITEMS":
            maybe_clean_new_items(final, req_type)
        cleaned = True

        log(
            f"✅ AMRO    | Subject: {subject}\n"
            f"           | Original: '{original_name}'\n"
            f"           | Type: '{req_type}'\n"
            f"           | Saved as: '{os.path.basename(final)}'\n"
            f"           | Folder: {dest_folder}"
            + (f"\n           | Cleaned: Yes" if cleaned else "")
        )

        show_notification(
            f"{label} ({sender_email})",
            subject,
            os.path.basename(final),
            dest_folder,
            cleaned=cleaned
        )

        _register_job(entry_id, sender_email, subject, os.path.basename(final), req_type)

        return True


# ──────────────────────────────────────────────────────────────
#  ASEEL EXCEL CLEANER
#  Applied to every file from aseel.d@example.com
#
#  Fixes done automatically:
#   1. Column C header "vendor" → "BARCODE"
#      (detected when column data is all long numeric strings ≥10 digits)
#   2. "Base UOM (PCS OR PCK)" column → replace any numeric cell with "PCS"
#   3. Add "YES" in a new column (header "YES") for every data row
#
#  Returns: (packing_reminder: bool)
#    True  → Base UOM had numeric values that were replaced → show reminder
#    False → Base UOM was fine
# ──────────────────────────────────────────────────────────────

def _looks_like_barcode_column(ws, col_idx: int, header_row: int) -> bool:
    """
    Return True if the data in this column looks like barcodes
    (≥80% of non-empty cells are numeric strings of ≥10 digits).
    """
    total = 0
    numeric_long = 0
    for row in ws.iter_rows(min_row=header_row + 1, max_row=header_row + 30,
                             min_col=col_idx, max_col=col_idx,
                             values_only=True):
        val = row[0]
        if val is None or str(val).strip() == "":
            continue
        total += 1
        if re.match(r'^\d{10,}$', str(val).strip()):
            numeric_long += 1
    return total > 0 and (numeric_long / total) >= 0.8


def clean_aseel_excel(filepath: str) -> bool:
    """
    Clean Aseel's Excel file in-place. Returns True if Base UOM had
    numeric packing values (packing reminder should be shown).
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        # ── Find header row (row 1 is usually the only header row) ──
        header_row = 1

        # Scan first 3 rows to find the one that has "ITEM NO" or similar
        for r in range(1, 4):
            row_vals = [str(ws.cell(row=r, column=c).value or "").upper()
                        for c in range(1, min(ws.max_column + 1, 30))]
            if any("ITEM" in v for v in row_vals):
                header_row = r
                break

        # Build col_map: normalised header → col index
        col_map = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=header_row, column=c).value
            if val:
                key = " ".join(str(val).upper().strip().split())
                col_map[key] = c

        packing_reminder = False
        # uom_rows: list of (item_no, pck_value, qty_number) captured BEFORE replacing with PCS
        uom_rows = []

        # ── Fix 1: Column C (or wherever header is "VENDOR") → "BARCODE" ──
        for header_text, col_idx in list(col_map.items()):
            if re.match(r'^VENDOR\s*$', header_text, re.IGNORECASE):
                if _looks_like_barcode_column(ws, col_idx, header_row):
                    ws.cell(row=header_row, column=col_idx).value = "BARCODE"
                    log(f"   🔧 Aseel fix: renamed column '{header_text}' → 'BARCODE' (col {col_idx})")
                    # Update col_map
                    col_map["BARCODE"] = col_idx
                    del col_map[header_text]
                    break

        # ── Fix 2: Base UOM column → replace numeric values with "PCS" ──
        uom_col = None
        for header_text, col_idx in col_map.items():
            if "BASE" in header_text and "UOM" in header_text:
                uom_col = col_idx
                break
        if uom_col is None:
            # Fallback: look for just "UOM"
            for header_text, col_idx in col_map.items():
                if "UOM" in header_text:
                    uom_col = col_idx
                    break

        if uom_col:
            # ── Capture ITEM NO column for UOM.txt ──
            item_no_col = None
            for h, c in col_map.items():
                if re.match(r'^ITEM(\s*(NO\.?|NUM(BER)?|#))?$', h, re.IGNORECASE):
                    item_no_col = c
                    break

            for row_idx in range(header_row + 1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=uom_col)
                val  = cell.value
                if val is None:
                    continue

                val_str = str(val).strip()

                # Detect packing quantity from multiple formats:
                #   Pure number:        6, 12, 15
                #   PCK-style:          PCK12, PCK 12
                #   PCK + newline PCS:  "PCK 12\n1"  (Aseel presses Enter then types 1)
                qty_str = ""

                if isinstance(val, (int, float)):
                    qty_str = str(int(float(val_str)))

                elif isinstance(val, str):
                    # Pure numeric string
                    if re.match(r'^\d+(\.\d+)?$', val_str):
                        qty_str = str(int(float(val_str)))
                    else:
                        # PCK12 / PCK 12 / PCK12\n1 / PCK 12\n1
                        m = re.match(r'^PCK\s*(\d+)', val_str, re.IGNORECASE)
                        if m:
                            qty_str = m.group(1)

                if qty_str:
                    # Get ITEM NO for this row
                    item_no_val = ""
                    if item_no_col:
                        iv = ws.cell(row=row_idx, column=item_no_col).value
                        item_no_val = str(iv).strip() if iv else ""
                    if item_no_val:
                        uom_rows.append((item_no_val, qty_str))
                    # Always replace cell with "PCS" regardless of original format
                    cell.value = "PCS"
                    packing_reminder = True

            if packing_reminder:
                log(f"   🔧 Aseel fix: PCK/numeric Base UOM values replaced with 'PCS' (col {uom_col})")

        # ── Fix 3: Add "YES" column at the end for every data row ──
        yes_col = ws.max_column + 1
        ws.cell(row=header_row, column=yes_col).value = "YES"
        rows_with_yes = 0
        for row_idx in range(header_row + 1, ws.max_row + 1):
            # Only write YES if there's any data in this row
            has_data = any(
                ws.cell(row=row_idx, column=c).value not in (None, "")
                for c in range(1, min(yes_col, 6))
            )
            if has_data:
                ws.cell(row=row_idx, column=yes_col).value = "YES"
                rows_with_yes += 1
        log(f"   🔧 Aseel fix: wrote 'YES' for {rows_with_yes} data row(s) (col {yes_col})")

        wb.save(filepath)
        log(f"   ✅ Aseel Excel cleaned: {os.path.basename(filepath)}")
        return packing_reminder, uom_rows

    except Exception as e:
        log(f"   ⚠️  clean_aseel_excel error: {e}")
        return False, []


UOM_TXT_FILE = r"C:\Users\abdul\Desktop\UOM.txt"

def write_uom_txt(uom_rows: list):
    """
    Write UOM packing data to UOM.txt on the Desktop.

    Navision dataport layout (tab-delimited, no header, Windows CRLF):
      col A  = ITEM NO                    e.g.  LOY032
      col B..G = empty (6 blank tabs)     skipped columns
      col H  = PCK+qty                    e.g.  PCK6
      col I  = qty number                 e.g.  6
      col J  = PCS                        always literal "PCS"
      (nothing after col J — no need to pad to col X)

    Appends to the file so multiple emails accumulate;
    a blank line separates batches.
    """
    if not uom_rows:
        return
    try:
        already_exists = os.path.isfile(UOM_TXT_FILE)
        lines = []
        if already_exists:
            lines.append("")   # blank separator between batches
        # 6 empty tabs = columns B, C, D, E, F, G
        GAP = "\t" * 6
        for item_no, qty_str in uom_rows:
            pck_val = f"PCK{qty_str}"        # e.g. PCK6, PCK12, PCK15
            lines.append(f"{item_no}\t{GAP}{pck_val}\t{qty_str}\tPCS")
        with open(UOM_TXT_FILE, "a", encoding="utf-8") as f:
            f.write("\r\n".join(lines) + "\r\n")
        log(f"   ✅ UOM.txt updated — {len(uom_rows)} row(s) written: {UOM_TXT_FILE}")
    except Exception as e:
        log(f"   ⚠️  write_uom_txt error: {e}")


def show_packing_reminder(sender_label: str, filename: str):
    """Show a separate persistent popup reminding to fill in packing quantities."""
    def _popup():
        root = tk.Tk()
        root.title("⚠️ Packing Reminder")
        root.attributes("-topmost", True)
        root.resizable(False, False)

        root.update_idletasks()
        sw, sh = root.winfo_screenwidth(), root.winfo_screenheight()
        w, h = 480, 200
        root.geometry(f"{w}x{h}+{sw - w - 20}+{sh - h - 320}")

        BG = "#1e1e2e"; ORANGE = "#fab387"; SUBTEXT = "#a6adc8"; TEXT = "#cdd6f4"

        root.configure(bg=BG)
        tk.Frame(root, bg=ORANGE, height=4).pack(fill="x", side="top")

        body = tk.Frame(root, bg=BG, padx=16, pady=14)
        body.pack(fill="both", expand=True)

        tf  = tkfont.Font(family="Segoe UI", size=11, weight="bold")
        nf  = tkfont.Font(family="Segoe UI", size=9)
        sf  = tkfont.Font(family="Segoe UI", size=8)

        tk.Label(body, text="⚠️  Packing Quantities — Action Required",
                 font=tf, bg=BG, fg=ORANGE, anchor="w").pack(fill="x")
        tk.Frame(body, bg=SUBTEXT, height=1).pack(fill="x", pady=(4, 10))

        tk.Label(body,
                 text=f"File:  {filename}\n\n"
                      f"Base UOM numbers (6, 12, 15…) were auto-set to 'PCS'.\n"
                      f"Please open the file and manually enter the correct\n"
                      f"packing quantities after item creation in Navision.",
                 font=nf, bg=BG, fg=TEXT, justify="left").pack(anchor="w")

        bf = tk.Frame(body, bg=BG)
        bf.pack(fill="x", pady=(10, 0))
        tk.Button(bf, text="Got it — I'll fill it in", command=root.destroy,
                  font=sf, bg=ORANGE, fg=BG, relief="flat",
                  padx=12, pady=4, cursor="hand2").pack(side="right")

        root.mainloop()

    threading.Thread(target=_popup, daemon=True).start()


def clean_aseel_change_price_columns(filepath: str):
    """
    For Aseel's CHANGE PRICE files:
      - Column B has photos but NO header text → delete it
      - Delete ANY column that has no header AND no data (all blank)
      - Keep columns that have either a header OR data values
    Saves the file in-place.

    From the screenshot layout:
      A = ITEM NO  (keep)
      B = photos, no header  (delete)
      C..N = empty, no header (delete)
      O, P = have data (keep)
      Q, R, S = have data (keep)
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        # Find header row (first row with recognisable content)
        header_row = 1
        for r in range(1, 4):
            row_vals = [str(ws.cell(row=r, column=c).value or "").strip()
                        for c in range(1, min(ws.max_column + 1, 30))]
            if any(v for v in row_vals):
                header_row = r
                break

        # Decide which columns to DELETE (right-to-left to avoid index shift)
        cols_to_delete = []
        for col_idx in range(1, ws.max_column + 1):
            header_val = str(ws.cell(row=header_row, column=col_idx).value or "").strip()

            # Check if any data exists in this column (skip header row)
            has_data = False
            for r in range(header_row + 1, min(ws.max_row + 1, header_row + 50)):
                cell_val = ws.cell(row=r, column=col_idx).value
                # Ignore image/None/empty — openpyxl reads images as None in value
                if cell_val is not None and str(cell_val).strip() != "":
                    has_data = True
                    break

            # Delete if: no header AND no data
            if not header_val and not has_data:
                cols_to_delete.append(col_idx)
                log(f"   🗑️  Aseel CP: deleting empty col {col_idx} (no header, no data)")

        # Delete right-to-left
        for col_idx in sorted(cols_to_delete, reverse=True):
            ws.delete_cols(col_idx)

        wb.save(filepath)
        log(f"   ✅ Aseel CP: removed {len(cols_to_delete)} empty/photo column(s)")

    except Exception as e:
        log(f"   ⚠️  clean_aseel_change_price_columns error: {e}")


def handle_aseel(att, sender_email: str, subject: str,
                 body: str, label: str, entry_id: str = "") -> bool:
    """
    Smart handler for Aseel:
      1. Detect type from subject/body (keyword scan) or Excel columns
      2. Apply Aseel-specific fixes:
           - vendor → BARCODE header fix
           - Base UOM numeric → PCS
           - Add YES column
      3. Save to correct folder with clean filename
      4. Show notification + packing reminder if needed
    """
    original_name = att.FileName
    ext = os.path.splitext(original_name)[1].lower()
    if ext not in ('.xlsx', '.xls', '.csv', '.txt'):
        return False

    today = datetime.now().strftime("%d-%m-%y")

    # ── Detect request type ──
    req_type, dest_folder = detect_type_from_text([subject, body])

    # Fallback: use Excel column scoring (same logic as Amro)
    tmp = _tmp_path(original_name)
    try:
        att.SaveAsFile(tmp)
    except Exception as e:
        log(f"   ⚠️  Aseel: could not save temp: {e}")
        return False

    if req_type is None:
        excel_type, _, _ = analyze_amro_excel(tmp)
        req_type    = excel_type or "NEW ITEMS"
        dest_folder = (CHANGE_PRICE_FOLDER
                       if req_type == "CHANGE PRICE"
                       else OPERATIONS_FOLDER)

    # ── Apply all Aseel fixes in-place on temp file ──
    packing_reminder, uom_rows = clean_aseel_excel(tmp)

    # ── Build filename ──
    # For CHANGE PRICE: read ITEM NO prefix from inside the Excel (e.g. W → "CHANGE PRICE W ASEEL 11-04-26")
    # For NEW ITEMS:    use supplier prefix from filename (e.g. LOY → "NEW ITEMS LOY ASEEL 08-04-26")
    if req_type == "CHANGE PRICE" and ext in ('.xlsx', '.xls'):
        # Clean the CHANGE PRICE file: remove photo column + all empty columns
        clean_aseel_change_price_columns(tmp)
        new_name = build_clean_filename(req_type, label, tmp, ext,
                                        original_name=original_name,
                                        subject=subject)
    else:
        # Extract supplier prefix from original filename
        _SKIP_WORDS = {
            "COPY", "OF", "THE", "A", "AN", "NEW", "IMPORT", "ITEMS", "ITEM",
            "CHANGE", "PRICE", "REDUCE", "UPLOAD", "FILE", "SHEET", "DATA", "LIST",
            "FINAL", "UPDATE", "REV", "REVISED", "APRIL", "MARCH", "MAY", "JUNE",
            "JULY", "AUGUST", "SEPTEMBER", "OCTOBER", "NOVEMBER", "DECEMBER",
            "JAN", "FEB", "MAR", "APR", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC",
        }
        stem = re.sub(r'[\-_\.]', ' ', os.path.splitext(original_name)[0])
        supplier_prefix = ""
        for word in stem.split():
            w = re.sub(r'[^A-Za-z]', '', word).upper()
            if len(w) >= 2 and w not in _SKIP_WORDS:
                supplier_prefix = w
                break
        log(f"   📋 Aseel supplier prefix: '{supplier_prefix}' (from '{original_name}')")
        parts = [req_type]
        if supplier_prefix:
            parts.append(supplier_prefix)
        parts.append(label)
        parts.append(today)
        new_name = " ".join(parts) + ext

    try:
        final = move_file_unique(tmp, dest_folder, new_name)
    except Exception as e:
        log(f"   ⚠️  Aseel: move failed ({e}); re-saving directly.")
        try:
            import shutil as _sh
            _sh.copy2(tmp, os.path.join(dest_folder, new_name))
            final = os.path.join(dest_folder, new_name)
        except Exception as e2:
            log(f"   ❌ Aseel: could not save: {e2}")
            return True

    log(
        f"✅ ASEEL   | Subject: {subject}\n"
        f"           | Original: '{original_name}'\n"
        f"           | Type: '{req_type}'\n"
        f"           | Saved as: '{os.path.basename(final)}'\n"
        f"           | Folder: {dest_folder}\n"
        f"           | Packing reminder: {packing_reminder}"
    )

    show_notification(f"{label} ({sender_email})", subject,
                      os.path.basename(final), dest_folder)

    _register_job(entry_id, sender_email, subject, os.path.basename(final), req_type)

    if packing_reminder:
        write_uom_txt(uom_rows)
        show_packing_reminder(label, os.path.basename(final))

    return True


# ──────────────────────────────────────────────────────────────
#  ZAINA HANDLER
#  Zaina sends NEW ITEMS files where the VENDOR NO is written
#  in the email body (e.g. "Kindly make creation for ART 20003")
#  rather than as a column in the Excel.
#
#  Logic:
#    1. Extract 5-6 digit vendor number from email body
#    2. Open the Excel — if VENDOR NO column is missing or all blank,
#       add / fill it with the number extracted from the body
#    3. Save to Operations with clean filename
# ──────────────────────────────────────────────────────────────

# All known VENDOR NO header variants (kept in sync with 03_import_textfiles.py)
_VENDOR_NO_HEADERS = {
    "VENDOR NO", "VENDOR NO.", "VENDOR NUMBER", "VENDOR NUMBERS",
    "VENDER NUMBER", "V.NO", "V.NO.", "SUPPLIER NO", "SUPPLIER NUMBER",
    "VENDOR ID", "VENDER", "COMPANY", "VENDORE COMP",
}

def extract_vendor_no_from_body(body: str) -> str:
    """
    Scan the email body for a standalone 5-6 digit number — that is
    the vendor number Zaina writes instead of putting it in the Excel.

    Examples:
      "Kindly make creation for ART 20003"  →  "20003"
      "Please create vendor 120456 items"   →  "120456"

    Rules:
      - Must be exactly 5 or 6 consecutive digits
      - Must NOT be part of a longer number (word-boundary anchored)
      - Returns the FIRST match found; empty string if nothing found
    """
    matches = re.findall(r'(?<!\d)(\d{5,6})(?!\d)', body)
    if matches:
        vendor_no = matches[0]
        log(f"   🔍 Zaina vendor NO extracted from body: '{vendor_no}'")
        return vendor_no
    log("   ⚠️  Zaina: no 5-6 digit vendor number found in email body.")
    return ""


def inject_vendor_no_into_excel(filepath: str, vendor_no: str) -> bool:
    """
    Open filepath, find the VENDOR NO column (any known alias).
    - If the column EXISTS but is blank/empty → fill every data row with vendor_no
    - If the column does NOT EXIST → add a new 'VENDOR NO' column at the end
      and fill every data row with vendor_no
    Returns True if anything was written, False otherwise.
    """
    if not vendor_no:
        return False
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        # Find header row (first row containing at least one recognisable header keyword)
        header_row = 1
        for r in range(1, 5):
            row_vals = [str(ws.cell(row=r, column=c).value or "").upper().strip()
                        for c in range(1, min(ws.max_column + 1, 50))]
            if any(v in _VENDOR_NO_HEADERS or "ITEM" in v or "BARCODE" in v
                   for v in row_vals):
                header_row = r
                break

        # Build col_map: normalised header → col index
        col_map = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=header_row, column=c).value
            if val:
                key = " ".join(str(val).upper().strip().split())
                col_map[key] = c

        # Find existing VENDOR NO column
        vendor_col = None
        for h, c in col_map.items():
            if h in _VENDOR_NO_HEADERS:
                vendor_col = c
                log(f"   📋 Zaina: found existing VENDOR NO column at col {c} (header='{h}')")
                break

        # Check if existing column already has real data
        if vendor_col:
            has_data = any(
                ws.cell(row=r, column=vendor_col).value not in (None, "")
                for r in range(header_row + 1, ws.max_row + 1)
            )
            if has_data:
                log(f"   ℹ️  Zaina: VENDOR NO column already has data — skipping inject.")
                wb.close()
                return False

        # If column missing → create it at the end
        if vendor_col is None:
            vendor_col = ws.max_column + 1
            ws.cell(row=header_row, column=vendor_col).value = "VENDOR NO"
            log(f"   🔧 Zaina: created new VENDOR NO column at col {vendor_col}")

        # Fill every data row that has an ITEM NO value
        item_no_col = None
        for h, c in col_map.items():
            if re.match(r'^ITEM(\s*(NO\.?|NUM(BER)?|#))?$', h, re.IGNORECASE):
                item_no_col = c
                break

        filled = 0
        for r in range(header_row + 1, ws.max_row + 1):
            # Only fill rows that have an item number (skip ghost/blank rows)
            if item_no_col:
                item_val = ws.cell(row=r, column=item_no_col).value
                if not item_val or str(item_val).strip() == "":
                    continue
            ws.cell(row=r, column=vendor_col).value = vendor_no
            filled += 1

        wb.save(filepath)
        log(f"   ✅ Zaina: wrote vendor NO '{vendor_no}' into {filled} row(s).")
        return filled > 0

    except Exception as e:
        log(f"   ⚠️  inject_vendor_no_into_excel error: {e}")
        return False


def _is_new_items_pricing(subject: str, body: str) -> bool:
    """
    Return True if this Zaina email is a NEW ITEMS PRICING scenario.
    Triggers when body/subject contains price/pricing keywords combined with
    new item keywords — but NOT a plain CHANGE PRICE email.

    Examples that match:
      body="TINA TOYS NEW ITEMS PRICING"
      body="New items price list"
      subject="New Item Pricing"
    """
    combined = (subject + " " + body).upper()
    has_price_kw  = bool(re.search(r'\bPRICING?\b', combined))
    has_new_item  = bool(re.search(r'\bNEW\s+ITEMS?\b', combined))
    # Exclude plain CHANGE PRICE emails (already handled by keyword detection)
    is_change_price = bool(re.search(r'CHANGE\s*PRICE|PRICE\s*CHANGE', combined))
    return has_price_kw and has_new_item and not is_change_price


# Known header aliases for ITEM NO and PRICE columns
_ITEM_NO_ALIASES = {
    "ITEM NO", "ITEMNO", "ITEMNO.", "ITEM", "ITEM NO.", "ITEM\nNO",
    "ITEM_NUM", "MARJI ITEM NO.", "ITEM NUMBER", "SKU", "NO.", "PRODUCT NO",
}
_PRICE_ALIASES = {
    "PRICE", "UNIT PRICE", "RSP", "RETAIL PRICE", "SELLING PRICE",
    "PRC", "SELLING\nPRICE", "SELLING\nPRICE 20%", "SALES PRICE",
    "MRP", "LIST PRICE", "S.PRISE", "CURRENT RSP", "CURRENT\nRSP",
}


def _normalise_header(val) -> str:
    """Normalise a cell value to uppercase stripped string."""
    return re.sub(r'\s+', ' ', str(val or "")).strip().upper()


def extract_items_and_prices(filepath: str) -> list:
    """
    Open an Excel file, find ITEM NO and PRICE columns (using broad alias sets),
    and return a list of (item_no, price) tuples — one per data row.
    Returns [] if columns not found or no data.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active

        # Find header row (first row that has recognisable content in col A)
        header_row = 1
        for r in range(1, 6):
            first_val = _normalise_header(ws.cell(row=r, column=1).value)
            if first_val:
                header_row = r
                break

        item_col  = None
        price_col = None
        for c in range(1, ws.max_column + 1):
            h = _normalise_header(ws.cell(row=header_row, column=c).value)
            if item_col  is None and h in _ITEM_NO_ALIASES:
                item_col  = c
                log(f"   📋 DMG: ITEM NO col={c} header='{h}'")
            if price_col is None and h in _PRICE_ALIASES:
                price_col = c
                log(f"   📋 DMG: PRICE col={c} header='{h}'")

        if not item_col or not price_col:
            log(f"   ⚠️  DMG: could not find ITEM NO (col={item_col}) "
                f"or PRICE (col={price_col}) — skipping DMG write.")
            wb.close()
            return []

        rows = []
        blank_streak = 0
        for r in range(header_row + 1, ws.max_row + 1):
            item_val  = ws.cell(row=r, column=item_col).value
            price_val = ws.cell(row=r, column=price_col).value
            if item_val is None and price_val is None:
                blank_streak += 1
                if blank_streak >= 10:
                    break
                continue
            blank_streak = 0
            item_str  = str(item_val).strip()  if item_val  is not None else ""
            price_str = str(price_val).strip() if price_val is not None else ""
            if item_str:
                rows.append((item_str, price_str))

        wb.close()
        log(f"   ✅ DMG: extracted {len(rows)} item/price row(s).")
        return rows

    except Exception as e:
        log(f"   ⚠️  extract_items_and_prices error: {e}")
        return []


def write_dmg_txt(rows: list) -> bool:
    """
    Write item/price rows to DMG.txt in Navision dataport format.
    ALWAYS overwrites the file with fresh data (no appending).

    Layout (tab-delimited, no header):
      col A = empty  (left blank)
      col B = ITEM NO
      col C = ALL
      col D = PCS
      col E = DATE   (DD-MM-YY)
      col F = PRICE

    Returns True on success.
    """
    if not rows:
        return False
    try:
        today = datetime.now().strftime("%d-%m-%y")
        lines = []
        for item_no, price in rows:
            lines.append(f"\t{item_no}\tALL\tPCS\t{today}\t{price}")
        with open(DMG_TXT_FILE, "w", encoding="utf-8") as f:   # "w" = overwrite
            f.write("\r\n".join(lines) + "\r\n")
        log(f"   ✅ DMG.txt written (overwrite) — {len(rows)} row(s): {DMG_TXT_FILE}")
        return True
    except Exception as e:
        log(f"   ❌ write_dmg_txt error: {e}")
        return False


def show_change_cost_reminder(sender_name: str = "The supplier"):
    """
    Non-blocking dark popup — reminds Abdul to update cost in Navision immediately
    when a CHANGE COST&PRICE file is received from any sender.
    Stays visible until manually dismissed.
    """
    def _popup():
        root = tk.Tk()
        root.title("⚡ Cost Update Required")
        root.attributes("-topmost", True)
        root.resizable(False, False)

        root.update_idletasks()
        sw, sh = root.winfo_screenwidth(), root.winfo_screenheight()
        w, h = 500, 270
        root.geometry(f"{w}x{h}+{(sw - w) // 2}+{(sh - h) // 2}")

        BG     = "#1e1e2e"
        RED    = "#f38ba8"
        ORANGE = "#fab387"
        TEAL   = "#89dceb"
        TEXT   = "#cdd6f4"
        SUB    = "#a6adc8"
        GREEN  = "#a6e3a1"

        root.configure(bg=BG)
        tk.Frame(root, bg=RED, height=5).pack(fill="x", side="top")

        body_frame = tk.Frame(root, bg=BG, padx=20, pady=16)
        body_frame.pack(fill="both", expand=True)

        tf  = tkfont.Font(family="Segoe UI", size=12, weight="bold")
        nf  = tkfont.Font(family="Segoe UI", size=10)
        sf  = tkfont.Font(family="Segoe UI", size=8)
        mf  = tkfont.Font(family="Segoe UI", size=11, weight="bold")

        tk.Label(body_frame, text="⚡  Action Required — Cost & Price Update",
                 font=tf, bg=BG, fg=RED, anchor="w").pack(fill="x")
        tk.Frame(body_frame, bg=SUB, height=1).pack(fill="x", pady=(5, 12))

        tk.Label(body_frame,
                 text=f"{sender_name} has sent a CHANGE COST & PRICE file.",
                 font=nf, bg=BG, fg=TEXT, anchor="w").pack(fill="x")

        tk.Frame(body_frame, bg=BG, height=6).pack()

        tk.Label(body_frame,
                 text="⚠️  Do not forget to update the COST in Navision immediately\n"
                      "    before processing the price change — cost must always\n"
                      "    be entered first so margins stay accurate.",
                 font=nf, bg=BG, fg=ORANGE, justify="left").pack(anchor="w")

        tk.Frame(body_frame, bg=BG, height=8).pack()

        tk.Label(body_frame,
                 text="📋  Steps:   1 → Update Cost in Navision   "
                      "2 → Then process Price change",
                 font=mf, bg=BG, fg=TEAL, anchor="w").pack(fill="x")

        tk.Frame(body_frame, bg=SUB, height=1).pack(fill="x", pady=(12, 6))

        bf = tk.Frame(body_frame, bg=BG)
        bf.pack(fill="x")
        tk.Button(bf, text="✅  Understood — I'll update Cost first",
                  font=sf, bg=RED, fg=BG, relief="flat",
                  padx=16, pady=6, cursor="hand2",
                  command=root.destroy).pack(side="right")

        root.mainloop()

    threading.Thread(target=_popup, daemon=True).start()


def fix_zaina_vendor_item_header(filepath: str):
    """
    Zeina-only fix applied before 03_import_textfiles.py reads the file.

    Problem:
      Zeina writes inconsistent headers for the VENDOR ITEM column:
        "VENDER NO.", "VENDOR NO.", "VENDOR ITEM NO.", "VENDOR" etc.
      These accidentally match the VENDOR NO alias list in 03_import_textfiles.py,
      causing vendor item codes (F1, 333-1, 22624 …) to land in the wrong column.

    Detection logic (two conditions BOTH must be true):
      1. File has a COMPANY column  → that is already mapped to VENDOR NO ✅
      2. File has another column whose header matches known Zeina VENDOR ITEM
         mislabels (e.g. "VENDOR NO.", "VENDER NO.", "VENDOR", "VENDOR ITEM NO.")

    Action:
      Rename that ambiguous column header to "VENDOR ITEM" so Script 2
      maps it correctly — without touching any cell data.

    Safe guard:
      If COMPANY column is missing, do nothing — we cannot be sure which
      column is VENDOR NO and which is VENDOR ITEM.
    """
    # Headers Zeina uses for what is actually the VENDOR ITEM column
    ZEINA_VENDOR_ITEM_MISLABELS = {
        "VENDOR NO.", "VENDER NO.", "VENDOR NO",
        "VENDOR ITEM NO.", "VENDOR ITEM NO",
        "VENDOR", "VENDER",
    }
    # Header that confirms VENDOR NO is already covered by COMPANY column
    COMPANY_VARIANTS = {"COMPANY", "COMPANY NAME"}

    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        # Find header row (first row with any content)
        header_row = 1
        for r in range(1, 5):
            row_vals = [str(ws.cell(row=r, column=c).value or "").strip()
                        for c in range(1, min(ws.max_column + 1, 30))]
            if any(v for v in row_vals):
                header_row = r
                break

        # Map normalised header → col index
        col_map = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=header_row, column=c).value
            if val:
                key = " ".join(str(val).strip().upper().split())
                col_map[key] = c

        # Condition 1: COMPANY column must exist
        has_company = any(k in COMPANY_VARIANTS for k in col_map)
        if not has_company:
            log("   ℹ️  Zaina header fix: no COMPANY column found — skipping.")
            return

        # Condition 2: find the mislabelled VENDOR ITEM column
        target_col = None
        original_header = ""
        for key, col_idx in col_map.items():
            if key in ZEINA_VENDOR_ITEM_MISLABELS:
                target_col = col_idx
                original_header = ws.cell(row=header_row, column=col_idx).value
                break

        if target_col is None:
            log("   ℹ️  Zaina header fix: no mislabelled VENDOR ITEM column found — skipping.")
            return

        # Rename the header
        ws.cell(row=header_row, column=target_col).value = "VENDOR ITEM"
        wb.save(filepath)
        log(f"   ✅ Zaina header fix: renamed col {target_col} "
            f"'{original_header}' → 'VENDOR ITEM'")

    except Exception as e:
        log(f"   ⚠️  fix_zaina_vendor_item_header error: {e}")


def handle_zaina(att, sender_email: str, subject: str,
                 body: str, label: str, entry_id: str = "") -> bool:
    """
    Handler for Zaina (zaina.n@example.com):
      1. Detect request type from subject/body/filename
      2. Save to temp
      3. Extract vendor number from email body (5-6 digits)
      4. Inject vendor number into Excel if column is missing or blank
      5. Save to Operations with clean filename
    """
    original_name = att.FileName
    ext = os.path.splitext(original_name)[1].lower()
    if ext not in ('.xlsx', '.xls', '.csv', '.txt'):
        return False

    # ── NEW ITEMS PRICING scenario (body says "TINA TOYS NEW ITEMS PRICING" etc.) ──
    if _is_new_items_pricing(subject, body) and ext in ('.xlsx', '.xls'):
        tmp = _tmp_path(original_name)
        try:
            att.SaveAsFile(tmp)
        except Exception as e:
            log(f"   ⚠️  Zaina pricing: could not save temp: {e}")
            return False

        rows = extract_items_and_prices(tmp)
        dmg_written = write_dmg_txt(rows)

        # Save original Excel to Desktop with its original filename
        desktop_path = r"C:\Users\abdul\Desktop"
        excel_dest   = os.path.join(desktop_path, original_name)
        try:
            shutil.copy2(tmp, excel_dest)
            log(f"   ✅ Zaina pricing: Excel saved to Desktop: '{original_name}'")
        except Exception as e:
            log(f"   ⚠️  Zaina pricing: could not save Excel to Desktop: {e}")

        try:
            os.remove(tmp)
        except Exception:
            pass

        notify_msg = (f"✓ {len(rows)} row(s) → {os.path.basename(DMG_TXT_FILE)}"
                      if dmg_written else "⚠️ No item/price data found — DMG not written")
        log(
            f"✅ ZAINA   | Subject: {subject}\n"
            f"           | Type: NEW ITEMS PRICING\n"
            f"           | DMG rows: {len(rows)}\n"
            f"           | File: {DMG_TXT_FILE}"
        )
        show_notification(
            f"{label} ({sender_email})", subject,
            notify_msg, os.path.dirname(DMG_TXT_FILE),
            cleaned=False,
        )
        _register_job(entry_id, sender_email, subject,
                      os.path.basename(DMG_TXT_FILE), "NEW ITEMS PRICING")
        return True

    # ── Detect request type ──
    req_type, dest_folder = detect_type_from_text([subject, body, original_name])
    if req_type is None:
        req_type    = "NEW ITEMS"
        dest_folder = OPERATIONS_FOLDER

    # ── Save to temp ──
    tmp = _tmp_path(original_name)
    try:
        att.SaveAsFile(tmp)
    except Exception as e:
        log(f"   ⚠️  Zaina: could not save temp: {e}")
        return False

    # ── Extract vendor number from email body and inject into Excel ──
    vendor_no = extract_vendor_no_from_body(body)
    if vendor_no and ext in ('.xlsx', '.xls'):
        inject_vendor_no_into_excel(tmp, vendor_no)

    # ── Fix Zeina's mislabelled VENDOR ITEM column header ──
    if ext in ('.xlsx', '.xls'):
        fix_zaina_vendor_item_header(tmp)

    # ── Build filename ────────────────────────────────────────────────────────
    # Special case: if CHANGE PRICE AND the original filename/subject contains
    # "COST" (e.g. "CHANGE COST&PRICE ZAN 18-04-2026.xlsx") → use
    # "CHANGE COST&PRICE" as the type prefix so it's easy to spot in the folder.
    # All other logic (letter prefix, buyer, sender label, date) stays the same.
    if req_type == "CHANGE PRICE":
        combined_search = f"{original_name} {subject}".upper()
        has_cost = bool(re.search(r'\bCOST\b', combined_search))
        if has_cost:
            # Build manually with "CHANGE COST&PRICE" prefix
            today = datetime.now().strftime("%d-%m-%y")
            prefixes = extract_letter_prefix_from_excel(tmp, ext) if ext in ('.xlsx', '.xls') else []
            letter_prefix = build_prefix_segment(prefixes, label) if prefixes else ""
            buyer_name = extract_buyer_from_text(combined_search)
            parts = ["CHANGE COST&PRICE"]
            if letter_prefix:
                parts.append(letter_prefix)
            if buyer_name:
                parts.append(buyer_name)
            parts.append(label)
            parts.append(today)
            new_name = " ".join(parts) + ext
            log(f"   📝 Zaina: COST detected → filename '{new_name}'")
            show_change_cost_reminder(sender_name=label)
        else:
            new_name = build_clean_filename(req_type, label, tmp, ext,
                                            original_name=original_name,
                                            subject=subject)
    else:
        new_name = build_clean_filename(req_type, label, tmp, ext,
                                        original_name=original_name,
                                        subject=subject)

    try:
        final = move_file_unique(tmp, dest_folder, new_name)
    except Exception as e:
        log(f"   ⚠️  Zaina: move failed ({e}); re-saving directly.")
        try:
            import shutil as _sh
            _sh.copy2(tmp, os.path.join(dest_folder, new_name))
            final = os.path.join(dest_folder, new_name)
        except Exception as e2:
            log(f"   ❌ Zaina: could not save: {e2}")
            return True

    cleaned = maybe_clean_change_price(final, req_type)

    _register_job(entry_id, sender_email, subject, os.path.basename(final), req_type)

    log(
        f"✅ ZAINA   | Subject: {subject}\n"
        f"           | Original: '{original_name}'\n"
        f"           | Type: '{req_type}'\n"
        f"           | Vendor NO: '{vendor_no or 'not found'}'\n"
        f"           | Saved as: '{os.path.basename(final)}'\n"
        f"           | Folder: {dest_folder}"
        + (f"\n           | Cleaned: Yes" if cleaned else "")
    )
    show_notification(f"{label} ({sender_email})", subject,
                      os.path.basename(final), dest_folder, cleaned=cleaned)
    return True


# ──────────────────────────────────────────────────────────────
#  GENERIC SPECIAL SENDER HANDLER  (Hadeel, …)
# ──────────────────────────────────────────────────────────────

def handle_special_generic(att, sender_email: str, subject: str,
                            body: str, label: str, entry_id: str = "") -> bool:
    original_name = att.FileName
    ext = os.path.splitext(original_name)[1].lower()
    if ext not in ('.xlsx', '.xls', '.csv', '.txt'):
        return False

    today = datetime.now().strftime("%d-%m-%y")

    tmp = _tmp_path(original_name)
    try:
        att.SaveAsFile(tmp)
    except Exception as e:
        log(f"   ⚠️  Could not save temp: {e}")
        return False

    req_type, dest_folder = detect_type_from_text([subject, body])

    if req_type is None:
        headers = peek_excel_headers(tmp)
        req_type, dest_folder = detect_type_from_text(headers)

    if req_type:
        new_name = build_clean_filename(req_type, label, tmp, ext,
                                        original_name=original_name,
                                        subject=subject)
    else:
        new_name    = f"{label} {original_name}"
        dest_folder = OPERATIONS_FOLDER
        log(f"   ⚠️  No type detected for '{original_name}' from {label} — saved with label prefix.")

    try:
        final = move_file_unique(tmp, dest_folder, new_name)
    except Exception as e:
        log(f"   ⚠️  Move failed ({e}); re-saving directly.")
        try:
            att.SaveAsFile(os.path.join(dest_folder, new_name))
            final = os.path.join(dest_folder, new_name)
        except Exception as e2:
            log(f"   ❌ Could not save: {e2}")
            return True

    # ── Clean CHANGE PRICE file ──
    cleaned = maybe_clean_change_price(final, req_type)

    log(
        f"✅ SPECIAL | From: {label} <{sender_email}> | Subject: {subject}\n"
        f"           | Original: '{original_name}'\n"
        f"           | Detected: '{req_type or 'UNKNOWN'}'\n"
        f"           | Saved as: '{os.path.basename(final)}'\n"
        f"           | Folder: {dest_folder}"
        + (f"\n           | Cleaned: Yes" if cleaned else "")
    )
    show_notification(f"{label} ({sender_email})", subject,
                      os.path.basename(final), dest_folder, cleaned=cleaned)
    _register_job(entry_id, sender_email, subject,
                  os.path.basename(final), req_type or "NEW ITEMS")
    return True


# ──────────────────────────────────────────────────────────────
#  NORMAL SENDER HELPERS
# ──────────────────────────────────────────────────────────────

def detect_req_type_from_name(original_name: str) -> tuple:
    """
    Scan the original filename for a known alias.
    Returns (req_type, matched_alias) or (None, None).
    """
    name_upper = original_name.upper()
    for alias, correct in ALIAS_MAP.items():
        if alias.upper() in name_upper:
            return correct, alias
    return None, None


def get_sender_label(sender_email: str) -> str:
    """Return display name for a normal sender, or empty string."""
    return next(
        (name for email, name in SENDER_NAME_MAP.items()
         if email.lower() == sender_email.lower()),
        "",
    )


def extract_letter_prefix_from_excel(filepath: str, ext: str) -> list:
    """
    Read the ITEM NO column from an Excel file and return ALL unique letter prefixes.
    e.g.  all HT-xxxxx  →  ["HT"]        → single prefix → use as-is
          HT + AF mixed →  ["HT", "AF"]  → multiple → caller uses MIX counter
    Returns [] if nothing found.
    """
    if ext not in ('.xlsx', '.xls'):
        return []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active

        # Find ITEM NO column — covers all known header variants (No., SKU, ITEM NO, etc.)
        item_no_col = _find_item_no_col(ws)

        prefixes = []
        if item_no_col:
            for row in ws.iter_rows(min_row=2, max_row=500,
                                    min_col=item_no_col, max_col=item_no_col,
                                    values_only=True):
                val = row[0]
                if val in (None, "", "-"):
                    continue
                val = str(val).strip()
                m = re.match(r'^([A-Za-z]+)', val)
                if m:
                    p = m.group(1).upper()
                    if p not in prefixes:
                        prefixes.append(p)

        wb.close()
        log(f"   📋 Item prefix scan → col={item_no_col} prefixes={prefixes}")
        return prefixes

    except Exception as e:
        log(f"   ⚠️  extract_letter_prefix_from_excel error: {e}")
        return []


def extract_buyer_from_text(text: str) -> str:
    """
    Scan subject line or original filename for a buyer.
    Two passes:
      Pass 1 — numeric code  e.g. "101" → "IHAB"
      Pass 2 — buyer name directly  e.g. "IHAB" → "IHAB"
    Returns "" if no match.
    """
    text_upper = text.upper()

    # Pass 1: numeric code lookup  (e.g. "101" → "IHAB")
    for code, name in BUYER_CODE_MAP.items():
        if re.search(r'(?<!\d)' + re.escape(code) + r'(?!\d)', text_upper):
            return name

    # Pass 2: buyer name appears directly as a whole word
    # Build a set of all known buyer names from BUYER_CODE_MAP values
    known_names = set(BUYER_CODE_MAP.values())
    for name in known_names:
        if re.search(r'\b' + re.escape(name.upper()) + r'\b', text_upper):
            return name

    return ""



def build_clean_filename(req_type: str, sender_label: str,
                         filepath: str, ext: str,
                         original_name: str = "", subject: str = "") -> str:
    """
    Build a clean filename for any CHANGE PRICE file:
      CHANGE PRICE {W} {IHAB} {dd-mm-yy}.xlsx

    Sources:
      - Letter prefix (e.g. W, AF)  → read from ITEM NO column inside Excel
      - Buyer name (e.g. IHAB)      → detected from subject line or original filename
                                       via BUYER_CODE_MAP  (e.g. "101" → "IHAB")
      - sender_label                → from SENDER_NAME_MAP (optional, only if mapped)
    """
    today = datetime.now().strftime("%d-%m-%y")

    if req_type == "CHANGE PRICE" and ext in ('.xlsx', '.xls'):
        # 1. Letter prefixes list from inside Excel → build_prefix_segment handles MIX/MIX2
        prefixes      = extract_letter_prefix_from_excel(filepath, ext)
        letter_prefix = build_prefix_segment(prefixes, sender_label) if prefixes else ""

        # 2. Buyer name from subject or original filename
        search_text = f"{subject} {original_name}".upper()
        buyer_name  = extract_buyer_from_text(search_text)

        parts = [req_type]
        if letter_prefix:
            parts.append(letter_prefix)
        if buyer_name:
            parts.append(buyer_name)
        if sender_label:
            parts.append(sender_label)
        parts.append(today)
        log(f"   📝 Filename built: {' '.join(parts) + ext}  "
            f"(prefix='{letter_prefix}' buyer='{buyer_name}' sender='{sender_label}')")
        return " ".join(parts) + ext

    else:
        # NEW ITEMS and other types: also scan for buyer name
        parts = [req_type]

        # Try to get letter prefix from Excel (e.g. "LGT")
        if ext in ('.xlsx', '.xls'):
            ni_prefixes = extract_letter_prefix_from_excel(filepath, ext)
            if ni_prefixes and len(ni_prefixes) == 1:
                parts.append(ni_prefixes[0])
            elif ni_prefixes and len(ni_prefixes) > 1:
                key = f"{sender_label}-{today}"
                _mix_counter[key] = _mix_counter.get(key, 0) + 1
                parts.append(f"MIX{_mix_counter[key]}")

        # Try to get buyer name from subject/filename (e.g. "IHAB" from code "101")
        search_text = f"{subject} {original_name}".upper()
        buyer_name = extract_buyer_from_text(search_text)
        if buyer_name:
            parts.append(buyer_name)

        if sender_label:
            parts.append(sender_label)
        parts.append(today)
        return " ".join(parts) + ext


# ──────────────────────────────────────────────────────────────
#  INBOX PROCESSOR
# ──────────────────────────────────────────────────────────────

def _collect_messages(folder, depth: int = 0, max_depth: int = 1) -> list:
    """
    Collect unread messages that have attachments from a folder.

    KEY CHANGE: Uses Outlook's Restrict() filter instead of loading ALL items
    into memory with list(folder.Items). This prevents the "Out of memory or
    system resources" crash that occurs when the inbox is large.

    max_depth is 1 (Inbox + immediate subfolders only) — deeper recursion
    was the main cause of memory exhaustion.
    """
    result = []
    try:
        # ── Use Restrict to pre-filter on the Outlook side ──
        # This loads only unread messages into the Python process,
        # instead of pulling every email in the folder.
        items = folder.Items
        restricted = items.Restrict("[Unread] = True")
        for i in range(1, restricted.Count + 1):
            try:
                msg = restricted.Item(i)
                # Secondary guard: skip if no attachments (cheap check)
                if msg.Attachments.Count > 0:
                    result.append(msg)
            except Exception as e:
                log(f"   WARNING: Could not read message {i}: {e}")
    except Exception as e:
        log(f"   WARNING: Could not read folder items: {e}")

    if depth < max_depth:
        try:
            for i in range(1, folder.Folders.Count + 1):
                try:
                    sub = folder.Folders.Item(i)
                    result.extend(_collect_messages(sub, depth + 1, max_depth))
                except Exception as e:
                    log(f"   WARNING: Could not read subfolder {i}: {e}")
        except Exception as e:
            log(f"   WARNING: Could not iterate subfolders: {e}")

    return result


def process_inbox(ns, processed_ids: set) -> tuple:
    """Process unread emails once from Inbox and immediate subfolders.
    Returns (updated_processed_ids, count_processed)."""
    inbox    = ns.GetDefaultFolder(6)
    messages = _collect_messages(inbox)
    log(f"   📬 Found {len(messages)} unread message(s) with attachments to check.")

    count = 0

    for message in messages:
        try:
            entry_id = message.EntryID
            if entry_id in processed_ids:
                continue
            if not (message.UnRead and message.Attachments.Count > 0):
                continue

            subject = message.Subject or "(no subject)"
            sender  = (message.SenderEmailAddress or "unknown").lower()
            body    = message.Body or ""
            _current_entry_id   = message.EntryID
            _current_sender     = sender
            _current_subject    = subject

            for i in range(1, message.Attachments.Count + 1):
                att = message.Attachments.Item(i)

                # ── Skip non-data files (images, PDFs, signatures, etc.) ──
                att_ext = os.path.splitext(att.FileName)[1].lower()
                if att_ext not in ('.xlsx', '.xls', '.csv', '.txt'):
                    log(f"   ⏭️  Skipped non-data attachment: '{att.FileName}'")
                    continue

                # ── Block Purchase — check subject FIRST, any sender ──
                if _is_block_purchase_subject(subject):
                    handled = handle_block_purchase(att, sender, subject)
                    if handled:
                        count += 1
                        continue

                # ── Khozema — dedicated handler (Fix 1-5) ──
                if sender == KHOZEMA_EMAIL:
                    handled = handle_khozema(att, message, sender, subject, body, entry_id)
                    if handled:
                        count += 1
                        continue

                # ── Special senders ──
                if sender in SPECIAL_SENDERS:
                    info  = SPECIAL_SENDERS[sender]
                    label = info["label"]

                    if info.get("smart_amro") or info.get("smart_ahmad"):
                        handled = handle_amro(att, sender, subject, body, label, entry_id)

                    elif info.get("smart_aseel"):
                        handled = handle_aseel(att, sender, subject, body, label, entry_id)

                    elif info.get("smart_zaina"):
                        handled = handle_zaina(att, sender, subject, body, label, entry_id)

                    else:
                        handled = handle_special_generic(att, sender, subject, body, label, entry_id)


                    if handled:
                        count += 1
                        continue

                # ── Normal alias-based processing ──
                original_name        = att.FileName
                ext                  = os.path.splitext(original_name)[1].lower()
                req_type, matched_alias = detect_req_type_from_name(original_name)

                # Fallback: scan subject if filename had no keyword
                if not matched_alias:
                    req_type, _ = detect_type_from_text([subject, body])
                    matched_alias = req_type  # truthy if found

                if matched_alias:
                    dest_folder  = (CHANGE_PRICE_FOLDER
                                    if req_type == "CHANGE PRICE"
                                    else OPERATIONS_FOLDER)
                    sender_label = get_sender_label(sender)

                    # Save to temp so we can read item numbers before final save
                    tmp = _tmp_path(original_name)
                    try:
                        att.SaveAsFile(tmp)
                    except Exception as e:
                        log(f"   ⚠️  Could not save temp: {e}")
                        continue

                    new_name  = build_clean_filename(req_type, sender_label, tmp, ext,
                                                      original_name=original_name,
                                                      subject=subject)
                    try:
                        dest_path = move_file_unique(tmp, dest_folder, new_name)
                    except Exception as e:
                        log(f"   ⚠️  Move failed ({e}); re-saving directly.")
                        try:
                            att.SaveAsFile(os.path.join(dest_folder, new_name))
                            dest_path = os.path.join(dest_folder, new_name)
                        except Exception as e2:
                            log(f"   ❌ Could not save: {e2}")
                            continue

                    cleaned = maybe_clean_change_price(dest_path, req_type)

                    if req_type == "CHANGE PRICE":
                        combined_check = f"{original_name} {subject}".upper()
                        if re.search(r'\bCOST\b', combined_check):
                            sender_display = get_sender_label(sender) or sender.split("@")[0].split(".")[0].upper()
                            show_change_cost_reminder(sender_name=sender_display)

                    log(
                        f"✅ FIXED  | From: {sender} | Subject: {subject}\n"
                        f"         | Original: '{original_name}'\n"
                        f"         | Renamed:  '{new_name}'\n"
                        f"         | Saved to: {dest_path}"
                        + (f"\n         | Cleaned: Yes" if cleaned else "")
                    )
                    show_notification(sender, subject, new_name, dest_folder, cleaned=cleaned)
                    _register_job(_current_entry_id, _current_sender,
                                  _current_subject, new_name, req_type)
                    count += 1

                else:
                    if original_name.lower().endswith(('.xlsx', '.xls', '.csv', '.txt')):
                        dest_path = save_attachment_direct(att, OPERATIONS_FOLDER, original_name)
                        log(
                            f"ℹ️  SAVED  | From: {sender} | Subject: {subject}\n"
                            f"         | File: '{original_name}' (no rename)\n"
                            f"         | Saved to: {dest_path}"
                        )
                        show_notification(sender, subject, original_name, OPERATIONS_FOLDER)
                        _register_job(_current_entry_id, _current_sender,
                                      _current_subject, original_name, "NEW ITEMS")
                        count += 1

            processed_ids.add(entry_id)

        except Exception as e:
            log(f"⚠️  ERROR processing message: {e}")
            continue

    # ── Post-processing: archive Buyers folder after successful processing ──
    # if count > 0:
    #    archive_buyers_files(OPERATIONS_FOLDER)

    return processed_ids, count


# ──────────────────────────────────────────────────────────────
#  JOB REGISTRY  — records every successfully processed email
# ──────────────────────────────────────────────────────────────

def _register_job(entry_id: str, sender_email: str, subject: str,
                  saved_filename: str, req_type: str):
    """
    Called once per saved file. Stores metadata so the
    'Done — Send Reply' button can find the right email later.
    """
    # Look up sender display name (label) from maps
    sender_name = ""
    if sender_email in SPECIAL_SENDERS:
        sender_name = SPECIAL_SENDERS[sender_email]["label"].capitalize()
    elif sender_email in SENDER_NAME_MAP:
        sender_name = SENDER_NAME_MAP[sender_email].capitalize()
    else:
        # Derive a name from the email local part e.g. "hussain.n" → "Hussain"
        local = sender_email.split("@")[0].split(".")[0]
        sender_name = local.capitalize()

    _processed_jobs.append({
        "entry_id":       entry_id,
        "sender_email":   sender_email,
        "sender_name":    sender_name,
        "subject":        subject,
        "saved_filename": saved_filename,
        "req_type":       req_type,
        "processing_script": "03_import_textfiles.py" if req_type == "NEW ITEMS" else "change_price-buyers.py",
        "replied":        False,
    })
    log(f"   📋 Job registered: '{saved_filename}' from {sender_email} ({sender_name})")


def _get_first_name(sender_email: str) -> str:
    """
    Returns the buyer's first name for the 'Dear X,' salutation.
    Priority:
      1. SPECIAL_SENDERS label (e.g. ZEINA → Zeina)
      2. SENDER_NAME_MAP       (e.g. AHMAD → Ahmad)
      3. Email local part      (e.g. hussain.n@... → Hussain)
    """
    if sender_email in SPECIAL_SENDERS:
        return SPECIAL_SENDERS[sender_email]["label"].capitalize()
    if sender_email in SENDER_NAME_MAP:
        return SENDER_NAME_MAP[sender_email].capitalize()
    local = sender_email.split("@")[0].split(".")[0]
    return local.capitalize()


def open_completion_reply(job: dict):
    """
    Opens a pre-filled Outlook reply window for the buyer to review before sending.
    - Replies to the original email thread (same EntryID)
    - CC always: ayman.s@example.com
    - Body: Dear [Name], \n The new items creation has been completed.
    - Opens in Outlook compose — user hits Send themselves

    Must be called from a thread that has called pythoncom.CoInitialize().
    """
    try:
        import pythoncom
        pythoncom.CoInitialize()

        outlook  = win32com.client.Dispatch("Outlook.Application")
        ns       = outlook.GetNamespace("MAPI")

        # Retrieve the original email by EntryID
        original = ns.GetItemFromID(job["entry_id"])

        # Build reply
        reply    = original.Reply()

        # Add CC — ayman always; Hussain for Zaina; Hadeel + Ayman for Ahmad
        zaina_email = "zaina.n@example.com"
        hussain_email = "HUSSAIN.N@example.com"
        ahmad_email = "ahmad.w@example.com"
        hadeel_email = "Hadeel.c@example.com"
        ayman_email = "ayman.s@example.com"

        effective_cc = REPLY_CC  # Default: ayman.s@example.com

        if job.get("sender_email", "").lower() == zaina_email:
            effective_cc = REPLY_CC + "; " + hussain_email
        elif job.get("sender_email", "").lower() == ahmad_email:
            effective_cc = hadeel_email + "; " + ayman_email

        if reply.CC:
            if effective_cc not in reply.CC:
                reply.CC = reply.CC + "; " + effective_cc
        else:
            reply.CC = effective_cc

        # Build body — personalized greeting + file reference
        first_name = _get_first_name(job["sender_email"])

        # ── Custom messages for Zeina ──
        if job.get("sender_email", "").lower() == "zaina.n@example.com":
            if job.get("processing_script") == "change_price-buyers.py":
                completion_text = "Noted..."
            else:
                completion_text = "Done..."
        else:
            # ── Standard messages for all other users ──
            if job.get("processing_script") == "change_price-buyers.py":
                completion_text = "Noted, I will review and process this accordingly."
            elif job.get("req_type") == "NEW ITEMS PRICING":
                completion_text = "The new item prices have been entered."
            else:
                completion_text = "The new items creation has been completed."

        body_text = (
            f"Dear {first_name},\n\n"
            f"{completion_text}"
            f"{REPLY_SIGNATURE}"
        )
        # Prepend to existing quoted body so the thread is preserved
        try:
            # HTMLBody: wrap in a <pre> for plain text inside HTML reply
            existing_html = reply.HTMLBody or ""
            # ── Custom messages for Zeina ──
            if job.get("sender_email", "").lower() == "zaina.n@example.com":
                if job.get("processing_script") == "change_price-buyers.py":
                    completion_text_html = "Noted..."
                else:
                    completion_text_html = "Done..."
            else:
                # ── Standard messages for all other users ──
                if job.get("processing_script") == "change_price-buyers.py":
                    completion_text_html = "Noted, I will review and process this accordingly."
                elif job.get("req_type") == "NEW ITEMS PRICING":
                    completion_text_html = "The new item prices have been entered."
                else:
                    completion_text_html = "The new items creation has been completed."
            new_html = (
                f"<div style='font-family:Calibri,sans-serif;font-size:11pt'>"
                f"Dear {first_name},<br><br>"
                f"{completion_text_html}"
                f"<br><br>Thank you."
                f"<br><br>Best regards,<br>Abdul Khader"
                f"</div><br>"
            ) + existing_html
            reply.HTMLBody = new_html
        except Exception:
            reply.Body = body_text + "\n\n" + (reply.Body or "")

        # Display — opens compose window for review, does NOT auto-send
        reply.Display(True)   # True = modal (waits until user closes/sends)

        # Mark as replied in session store
        job["replied"] = True
        log(f"   ✅ Reply opened in Outlook for job: '{job['saved_filename']}'")

    except Exception as e:
        log(f"   ⚠️  open_completion_reply error: {e}")
        # Show error in a simple messagebox so user is aware
        try:
            from tkinter import messagebox
            messagebox.showerror(
                "Mail2Nav — Reply Error",
                f"Could not open the reply in Outlook:\n\n{e}\n\n"
                f"Please reply manually to: {job['sender_email']}"
            )
        except Exception:
            pass
    finally:
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


# ──────────────────────────────────────────────────────────────
#  DONE — SEND REPLY PICKER  (shown when button is clicked)
# ──────────────────────────────────────────────────────────────

def show_reply_picker():
    """
    Opens a picker dialog listing all jobs from this session.
    User selects which one they have finished in Navision.
    Script then opens the Outlook reply for that job.
    """
    # Filter: only NEW ITEMS and CHANGE PRICE jobs not yet replied to
    pending = [j for j in _processed_jobs
               if j["req_type"] in ("NEW ITEMS", "CHANGE PRICE") and not j["replied"]]

    if not pending:
        all_jobs = [j for j in _processed_jobs if j["req_type"] in ("NEW ITEMS", "CHANGE PRICE")]
        if all_jobs:
            msg = "All NEW ITEMS and CHANGE PRICE emails from this session have already been replied to."
        else:
            msg = "No NEW ITEMS or CHANGE PRICE emails have been processed in this session yet.\n\nProcess emails first, then import into Navision, then click Done."
        from tkinter import messagebox
        messagebox.showinfo("Mail2Nav — No Pending Replies", msg)
        return

    # ── Build picker window ──────────────────────────────────
    win = tk.Toplevel()
    win.title("Mail2Nav — Done: Select Job")
    win.resizable(False, False)
    win.attributes("-topmost", True)

    BG      = "#1e1e2e"
    ACCENT  = "#a6e3a1"   # green for "done" action
    TEXT    = "#cdd6f4"
    SUBTEXT = "#a6adc8"
    ORANGE  = "#fab387"

    win.configure(bg=BG)
    win.update_idletasks()
    sw, sh = win.winfo_screenwidth(), win.winfo_screenheight()
    w, h   = 500, min(160 + len(pending) * 68, 520)
    win.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

    tk.Frame(win, bg=ACCENT, height=4).pack(fill="x", side="top")

    # Header
    hf = tk.Frame(win, bg=BG, pady=14, padx=20)
    hf.pack(fill="x")
    tf = tkfont.Font(family="Segoe UI", size=13, weight="bold")
    tk.Label(hf, text="📬  Mail2Nav",
             font=tf, bg=BG, fg=ACCENT).pack(anchor="w")
    nf = tkfont.Font(family="Segoe UI", size=9)
    sf = tkfont.Font(family="Segoe UI", size=8)
    tk.Label(hf,
             text="Click Process when ready to import.\nAfter Navision creation, click Done — Send Reply.",
             font=nf, bg=BG, fg=SUBTEXT, justify="left").pack(anchor="w", pady=(4, 0))

    tk.Frame(win, bg=SUBTEXT, height=1).pack(fill="x", padx=20)

    # ── One card per pending job ──
    scroll_frame = tk.Frame(win, bg=BG)
    scroll_frame.pack(fill="both", expand=True)

    def _pick(job, w=win):
        w.destroy()
        threading.Thread(
            target=open_completion_reply, args=(job,), daemon=True
        ).start()

    def _skip(job, card_frame):
        """Mark job as replied (dismissed) and remove its card from the picker."""
        job["replied"] = True
        card_frame.destroy()
        log(f"   ⏭️  Job skipped (no reply): '{job['saved_filename']}'")

    for job in pending:
        card = tk.Frame(scroll_frame, bg="#2a2a3e", padx=12, pady=8,
                        highlightbackground="#3a3a5e", highlightthickness=1)
        card.pack(fill="x", pady=4)

        left  = tk.Frame(card, bg="#2a2a3e")
        left.pack(side="left", fill="both", expand=True)

        tk.Label(left,
                 text=f"To: {job['sender_name']}  ({job['sender_email']})",
                 font=nf, bg="#2a2a3e", fg=TEXT, anchor="w").pack(fill="x")
        tk.Label(left,
                 text=f"File: {job['saved_filename']}",
                 font=sf, bg="#2a2a3e", fg=SUBTEXT, anchor="w").pack(fill="x")
        tk.Label(left,
                 text=f"Subject: {job['subject'][:55]}{'…' if len(job['subject'])>55 else ''}",
                 font=sf, bg="#2a2a3e", fg=SUBTEXT, anchor="w").pack(fill="x")

        btn_frame = tk.Frame(card, bg="#2a2a3e")
        btn_frame.pack(side="right", padx=(8, 0))

        tk.Button(btn_frame,
                  text="Open Reply ▶",
                  font=sf, bg=ACCENT, fg=BG,
                  relief="flat", padx=10, pady=5,
                  cursor="hand2",
                  command=lambda j=job: _pick(j)).pack(side="top", pady=(0, 4))

        tk.Button(btn_frame,
                  text="Skip — No Reply",
                  font=sf, bg="#45475a", fg=TEXT,
                  relief="flat", padx=10, pady=4,
                  cursor="hand2",
                  command=lambda j=job, c=card: _skip(j, c)).pack(side="top")

    tk.Frame(win, bg=SUBTEXT, height=1).pack(fill="x", pady=(10, 4))
    tk.Label(win,
             text=f"Reply will include CC: {REPLY_CC}",
             font=sf, bg=BG, fg=ORANGE, anchor="w").pack(fill="x")


# ──────────────────────────────────────────────────────────────
#  MAIN CONTROL PANEL (Tkinter)
# ──────────────────────────────────────────────────────────────

class ControlPanel:
    """
    Main window with a single 'Process New Emails' button.
    Processing runs in a background thread so the UI stays responsive.
    """

    BG       = "#1e1e2e"
    ACCENT   = "#89b4fa"
    TEXT     = "#cdd6f4"
    SUBTEXT  = "#a6adc8"
    GREEN    = "#a6e3a1"
    RED      = "#f38ba8"
    YELLOW   = "#f9e2af"

    def __init__(self):
        self.processed_ids = set()
        self.processing    = False

        self.root = tk.Tk()
        self.root.title("📬 Mail2Nav")
        self.root.resizable(True, True)
        self.root.configure(bg=self.BG)
        self.root.attributes("-topmost", True)
        self.root.minsize(380, 300)

        # Position: right side of screen, comfortable height
        self.root.update_idletasks()
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        w, h = 420, 340
        x = sw - w - 20          # 20px from right edge
        y = max(10, (sh - h) // 2 - 60)   # shifted upward so taskbar stays visible
        self.root.geometry(f"{w}x{h}+{x}+{y}")

        self._build_ui()
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

    # ── UI construction ──────────────────────────────────────

    def _build_ui(self):
        BG      = self.BG
        ACCENT  = self.ACCENT
        TEXT    = self.TEXT
        SUBTEXT = self.SUBTEXT

        # Top accent bar
        tk.Frame(self.root, bg=ACCENT, height=4).pack(fill="x", side="top")

        # Header
        hf = tk.Frame(self.root, bg=BG, pady=8, padx=20)
        hf.pack(fill="x")
        tf = tkfont.Font(family="Segoe UI", size=13, weight="bold")
        tk.Label(hf, text="📬  Mail2Nav",
                 font=tf, bg=BG, fg=ACCENT).pack(anchor="w")
        nf = tkfont.Font(family="Segoe UI", size=9)
        tk.Label(hf,
                 text="Click Process when ready to import.\nAfter Navision creation, click Done — Send Reply.",
                 font=nf, bg=BG, fg=SUBTEXT, justify="left").pack(anchor="w", pady=(4, 0))

        tk.Frame(self.root, bg=SUBTEXT, height=1).pack(fill="x", padx=20)

        # Main button
        bf = tk.Frame(self.root, bg=BG, pady=10)
        bf.pack()
        btn_font = tkfont.Font(family="Segoe UI", size=12, weight="bold")
        self.btn = tk.Button(
            bf,
            text="▶  Process New Emails",
            font=btn_font,
            bg=ACCENT,
            fg=BG,
            activebackground="#6fa8e8",
            activeforeground=BG,
            relief="flat",
            padx=28, pady=12,
            cursor="hand2",
            command=self._on_process_click,
        )
        self.btn.pack()

        # ── Done — Send Reply button ──────────────────────────
        ORANGE = "#fab387"
        self.reply_btn = tk.Button(
            bf,
            text="✉  Done — Send Reply",
            font=tkfont.Font(family="Segoe UI", size=10),
            bg=ORANGE,
            fg=BG,
            activebackground="#d4845a",
            activeforeground=BG,
            relief="flat",
            padx=18, pady=8,
            cursor="hand2",
            command=self._on_reply_click,
        )
        self.reply_btn.pack(pady=(10, 0))

        # Status label
        sf = tkfont.Font(family="Segoe UI", size=9)
        self.status_var = tk.StringVar(value="⏸  Idle — waiting for you to click.")
        self.status_lbl = tk.Label(
            self.root,
            textvariable=self.status_var,
            font=sf, bg=BG, fg=SUBTEXT,
        )
        self.status_lbl.pack(pady=(0, 6))

        tk.Frame(self.root, bg=SUBTEXT, height=1).pack(fill="x", padx=20)

        # Info rows
        info_frame = tk.Frame(self.root, bg=BG, padx=20, pady=10)
        info_frame.pack(fill="x")
        lf = tkfont.Font(family="Segoe UI", size=8)

        def info_row(label_text, value_text, fg=SUBTEXT):
            f = tk.Frame(info_frame, bg=BG)
            f.pack(fill="x", pady=1)
            tk.Label(f, text=label_text, width=14, anchor="w",
                     font=lf, bg=BG, fg=SUBTEXT).pack(side="left")
            tk.Label(f, text=value_text, anchor="w",
                     font=lf, bg=BG, fg=fg).pack(side="left")

        info_row("Operations:",   OPERATIONS_FOLDER,   fg=TEXT)
        info_row("Change Price:", CHANGE_PRICE_FOLDER, fg=TEXT)
        info_row("Special senders:", ", ".join(
            v["label"] for v in SPECIAL_SENDERS.values()), fg=TEXT)
        info_row("Auto-cleanup:", "Buyer · Arabic Desc · Vendor No.", fg=self.YELLOW)

        # Last-run label
        lrf = tkfont.Font(family="Segoe UI", size=8)
        self.last_run_var = tk.StringVar(value="Last run: never")
        tk.Label(self.root, textvariable=self.last_run_var,
                 font=lrf, bg=BG, fg=SUBTEXT).pack(pady=(4, 10))

    # ── Button handler ───────────────────────────────────────

    def _on_reply_click(self):
        """Open the reply picker on the main thread (Tkinter requires it)."""
        self.root.after(0, show_reply_picker)

    def _on_process_click(self):
        if self.processing:
            return
        self.processing = True
        self.btn.config(state="disabled", text="⏳  Processing…", bg="#555577")
        self.status_var.set("🔄  Scanning inbox…")
        self.root.update()
        threading.Thread(target=self._run_processing, daemon=True).start()

    def _run_processing(self):
        """
        Runs in a background thread.
        win32com COM objects are apartment-threaded and CANNOT be shared across
        threads - so we create a fresh Outlook connection here each time.
        """
        try:
            log("\n" + "─" * 50)
            log("▶  Manual processing triggered by user")

            # Re-initialize COM in this thread
            import pythoncom
            pythoncom.CoInitialize()
            try:
                outlook = win32com.client.Dispatch("Outlook.Application")
                ns      = outlook.GetNamespace("MAPI")
                self.processed_ids, count = process_inbox(ns, self.processed_ids)
            finally:
                pythoncom.CoUninitialize()

            log(f"✔  Done. {count} new email(s) processed.")
            msg = (f"✅  Done — {count} email(s) processed."
                   if count > 0
                   else "ℹ️   No new emails found.")
            self.root.after(0, self._processing_done, msg, count)
        except Exception as e:
            log(f"❌  Processing error: {e}")
            self.root.after(0, self._processing_done, f"❌  Error: {e}", 0)

    def _processing_done(self, msg: str, count: int):
        now = datetime.now().strftime("%H:%M:%S")
        self.last_run_var.set(f"Last run: {now}  ({count} processed)")
        self.status_var.set(msg)
        fg = self.GREEN if count > 0 else self.YELLOW
        self.status_lbl.config(fg=fg)
        self.btn.config(state="normal", text="▶  Process New Emails", bg=self.ACCENT)
        self.processing = False

    # ── Lifecycle ────────────────────────────────────────────

    def _on_close(self):
        log("🛑 Email Attachment Fixer — closed by user.")
        self.root.destroy()

    def run(self):
        self.root.mainloop()


# ──────────────────────────────────────────────────────────────
#  ENTRY POINT
# ──────────────────────────────────────────────────────────────

def main():
    log("=" * 60)
    log("Mail2Nav v3 — STARTED (Manual Mode)")
    log(f"Operations folder  : {OPERATIONS_FOLDER}")
    log(f"Change Price folder: {CHANGE_PRICE_FOLDER}")
    log(f"Mode               : Manual (button-triggered)")
    log(f"Alias rules        : {len(ALIAS_MAP)}")
    log(f"Special senders    : {', '.join(SPECIAL_SENDERS)}")
    log(f"Columns removed    : {', '.join(COLUMNS_TO_REMOVE)}")
    log("=" * 60)

    # Verify Outlook is reachable before showing the UI
    try:
        import pythoncom
        pythoncom.CoInitialize()
        outlook = win32com.client.Dispatch("Outlook.Application")
        outlook.GetNamespace("MAPI")   # quick connectivity check only
        pythoncom.CoUninitialize()
        log("✅ Outlook reachable. Launching control panel.")
    except Exception as e:
        log(f"❌ Failed to connect to Outlook: {e}")
        log("   Make sure Outlook is open and running.")
        root = tk.Tk()
        root.withdraw()
        from tkinter import messagebox
        messagebox.showerror(
            "Outlook Connection Failed",
            f"Could not connect to Outlook:\n\n{e}\n\nMake sure Outlook is open and try again."
        )
        root.destroy()
        return

    panel = ControlPanel()
    panel.run()


if __name__ == "__main__":
    main()